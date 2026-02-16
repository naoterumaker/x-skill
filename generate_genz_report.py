#!/usr/bin/env python3
"""
Z世代売れ筋アイテム X/Twitter トレンドレポート
5ジャンル横断分析 → xlsx出力
"""

import json, re, sys
from pathlib import Path
from datetime import datetime
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    from openpyxl.utils import get_column_letter

# --- Styling ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
BOLD = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
INSIGHT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
VIRAL_FILL = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
NUM_FMT = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

CATEGORY_STYLES = {
    "コスメ・美容": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "推し活グッズ": PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
    "ガジェット・QOL": PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
    "フード・スイーツ": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "ファッション": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
}

POST_TYPE_LABELS = {
    "quote": "引用", "x_article": "X記事", "article_link": "記事リンク",
    "media": "メディア", "text": "テキスト",
}

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_width(ws, min_w=8, max_w=60):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)

def compact(n):
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000: return f"{n/1_000:.1f}K"
    return str(int(n))

# --- Load ---
print("Loading data...", file=sys.stderr)
categories = {
    "コスメ・美容": "/tmp/genz-cosme.json",
    "推し活グッズ": "/tmp/genz-oshikatsu.json",
    "ガジェット・QOL": "/tmp/genz-gadget.json",
    "フード・スイーツ": "/tmp/genz-food.json",
    "ファッション": "/tmp/genz-fashion.json",
}

all_data = {}
seen = set()
for cat, path in categories.items():
    tweets = json.loads(Path(path).read_text())
    deduped = []
    for t in tweets:
        if t["id"] not in seen:
            seen.add(t["id"])
            t["category"] = cat
            deduped.append(t)
    all_data[cat] = deduped
    print(f"  {cat}: {len(deduped)} tweets", file=sys.stderr)

all_tweets = []
for tweets in all_data.values():
    all_tweets.extend(tweets)
all_tweets.sort(key=lambda t: t["metrics"]["likes"], reverse=True)
print(f"  Total: {len(all_tweets)} tweets", file=sys.stderr)

# --- Analysis helpers ---
def cat_stats(tweets):
    if not tweets: return {}
    likes = [t["metrics"]["likes"] for t in tweets]
    bmarks = [t["metrics"].get("bookmarks", 0) for t in tweets]
    return {
        "count": len(tweets),
        "total_likes": sum(likes),
        "avg_likes": sum(likes) / len(likes),
        "max_likes": max(likes),
        "total_bookmarks": sum(bmarks),
        "avg_save_rate": sum(bmarks) / sum(likes) if sum(likes) > 0 else 0,
    }

# --- Build Workbook ---
wb = Workbook()

# ===== Sheet 1: カテゴリ概要 =====
ws = wb.active
ws.title = "カテゴリ概要"
ws.cell(row=1, column=1, value="Z世代トレンドアイテム - カテゴリ別バズ分析").font = TITLE_FONT
ws.merge_cells("A1:G1")
ws.cell(row=2, column=1, value=f"分析日: {datetime.now().strftime('%Y-%m-%d')} | X直近7日間 | 合計{len(all_tweets)}件")

row = 4
ws.append(["カテゴリ", "件数", "平均いいね", "最大いいね", "平均保存率", "トップ投稿", "判定"])
style_header(ws, row, 7)

cat_rankings = []
for cat, tweets in all_data.items():
    s = cat_stats(tweets)
    if not s: continue
    top = max(tweets, key=lambda x: x["metrics"]["likes"])
    cat_rankings.append((cat, s, top))

cat_rankings.sort(key=lambda x: x[1]["avg_likes"], reverse=True)

for cat, s, top in cat_rankings:
    row += 1
    # Determine hot level
    if s["avg_likes"] > 3000:
        verdict = "激アツ"
    elif s["avg_likes"] > 1000:
        verdict = "アツい"
    else:
        verdict = "注目"
    ws.cell(row=row, column=1, value=cat)
    ws.cell(row=row, column=1).font = BOLD
    fill = CATEGORY_STYLES.get(cat)
    if fill:
        ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=2, value=s["count"])
    ws.cell(row=row, column=3, value=int(s["avg_likes"]))
    ws.cell(row=row, column=3).number_format = NUM_FMT
    ws.cell(row=row, column=4, value=s["max_likes"])
    ws.cell(row=row, column=4).number_format = NUM_FMT
    ws.cell(row=row, column=5, value=f"{s['avg_save_rate']:.1%}")
    ws.cell(row=row, column=6, value=f"@{top['username']}: {top['text'][:100]}")
    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
    ws.cell(row=row, column=7, value=verdict)
    ws.cell(row=row, column=7).font = BOLD
    if verdict == "激アツ":
        ws.cell(row=row, column=7).fill = VIRAL_FILL

ws.column_dimensions["A"].width = 18
ws.column_dimensions["F"].width = 70
ws.column_dimensions["G"].width = 10
ws.freeze_panes = "A5"

# ===== Sheet 2: 全投稿ランキング（バズ効率順） =====
ws2 = wb.create_sheet("バズ効率ランキング")
ws2.cell(row=1, column=1, value="バズ効率TOP（フォロワー対比で異常にバズった投稿）").font = TITLE_FONT
ws2.merge_cells("A1:H1")

row = 3
headers = ["No", "カテゴリ", "Username", "アカウントURL", "テキスト", "ポストURL",
           "いいね", "フォロワー", "バズ効率", "保存率", "投稿タイプ"]
ws2.append(headers)
style_header(ws2, row, len(headers))

eff_sorted = sorted(all_tweets, key=lambda t: t["metrics"]["likes"] / max(t.get("author_followers", 1), 1), reverse=True)
for i, t in enumerate(eff_sorted[:50], 1):
    row += 1
    followers = t.get("author_followers", 1) or 1
    likes = t["metrics"]["likes"]
    bmarks = t["metrics"].get("bookmarks", 0)
    eff = likes / followers
    save_rate = bmarks / likes if likes > 0 else 0
    cat = t.get("category", "?")
    ws2.cell(row=row, column=1, value=i)
    ws2.cell(row=row, column=2, value=cat)
    fill = CATEGORY_STYLES.get(cat)
    if fill:
        ws2.cell(row=row, column=2).fill = fill
    ws2.cell(row=row, column=3, value=f"@{t['username']}")
    ws2.cell(row=row, column=4, value=t.get("account_url", f"https://x.com/{t['username']}"))
    ws2.cell(row=row, column=5, value=t["text"][:200])
    ws2.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
    ws2.cell(row=row, column=6, value=t.get("tweet_url", ""))
    ws2.cell(row=row, column=7, value=likes)
    ws2.cell(row=row, column=7).number_format = NUM_FMT
    ws2.cell(row=row, column=8, value=followers)
    ws2.cell(row=row, column=8).number_format = NUM_FMT
    ws2.cell(row=row, column=9, value=round(eff, 2))
    ws2.cell(row=row, column=10, value=f"{save_rate:.1%}")
    ws2.cell(row=row, column=11, value=POST_TYPE_LABELS.get(t.get("post_type", "text"), t.get("post_type", "")))
    if eff >= 1.0:
        ws2.cell(row=row, column=9).fill = VIRAL_FILL
        ws2.cell(row=row, column=9).font = BOLD

auto_width(ws2, max_w=50)
ws2.column_dimensions["D"].width = 30
ws2.column_dimensions["E"].width = 60
ws2.column_dimensions["F"].width = 45
ws2.freeze_panes = "A4"

# ===== Sheet 3-7: カテゴリ別詳細 =====
for cat, tweets in all_data.items():
    ws_cat = wb.create_sheet(cat)
    headers = ["No", "Username", "アカウントURL", "テキスト", "ポストURL",
               "いいね", "RT", "ブックマーク", "バズ効率", "保存率",
               "投稿タイプ", "投稿日時"]
    ws_cat.append(headers)
    style_header(ws_cat, 1, len(headers))

    for i, t in enumerate(sorted(tweets, key=lambda x: x["metrics"]["likes"], reverse=True), 1):
        m = t["metrics"]
        followers = t.get("author_followers", 1) or 1
        likes = m["likes"]
        bmarks = m.get("bookmarks", 0)
        eff = likes / followers
        save_rate = bmarks / likes if likes > 0 else 0
        ws_cat.append([
            i, f"@{t['username']}",
            t.get("account_url", f"https://x.com/{t['username']}"),
            t["text"][:300],
            t.get("tweet_url", ""),
            likes, m["retweets"], bmarks,
            round(eff, 2), f"{save_rate:.1%}",
            POST_TYPE_LABELS.get(t.get("post_type", "text"), ""),
            t.get("created_at", ""),
        ])
        row_num = ws_cat.max_row
        for c in [6, 7, 8]:
            ws_cat.cell(row=row_num, column=c).number_format = NUM_FMT
        if eff >= 1.0:
            ws_cat.cell(row=row_num, column=9).fill = VIRAL_FILL
            ws_cat.cell(row=row_num, column=9).font = BOLD

    auto_width(ws_cat, max_w=50)
    ws_cat.column_dimensions["C"].width = 30
    ws_cat.column_dimensions["D"].width = 60
    ws_cat.column_dimensions["E"].width = 45
    ws_cat.freeze_panes = "A2"

# ===== Sheet 8: 戦略的インサイト =====
ws_ins = wb.create_sheet("戦略的インサイト")
ws_ins.cell(row=1, column=1, value="Z世代向けアイテム - 戦略的インサイト").font = TITLE_FONT
ws_ins.merge_cells("A1:C1")

insights = []

# 1. カテゴリ比較
cat_rankings.sort(key=lambda x: x[1]["avg_likes"], reverse=True)
best_cat = cat_rankings[0]
insights.append({
    "priority": "HIGH",
    "category": "最もバズるジャンル",
    "insight": f"「{best_cat[0]}」が平均{compact(best_cat[1]['avg_likes'])}いいねで最強。最大{compact(best_cat[1]['max_likes'])}いいね。"
})

# 2. Save rate comparison
save_ranked = sorted(cat_rankings, key=lambda x: x[1]["avg_save_rate"], reverse=True)
best_save = save_ranked[0]
insights.append({
    "priority": "HIGH",
    "category": "最もストック型（保存率が高い）",
    "insight": f"「{best_save[0]}」が保存率{best_save[1]['avg_save_rate']:.1%}でトップ。ブクマされる=何度も見返される=購買に繋がりやすい。"
})

# 3. Viral efficiency heroes
top_eff = sorted(all_tweets, key=lambda t: t["metrics"]["likes"] / max(t.get("author_followers", 1), 1), reverse=True)[:5]
for t in top_eff[:3]:
    followers = t.get("author_followers", 1) or 1
    eff = t["metrics"]["likes"] / followers
    insights.append({
        "priority": "HIGH",
        "category": f"バズ効率の怪物（{t.get('category', '?')}）",
        "insight": f"@{t['username']}（フォロワー{compact(followers)}）が{compact(t['metrics']['likes'])}いいね獲得（効率{eff:.0f}倍）。内容: {t['text'][:80]}"
    })

# 4. Product extraction from text
product_keywords = {}
for t in all_tweets:
    text = t["text"]
    # Extract potential product mentions
    for pattern in [
        r'「([^」]{2,20})」',  # 「商品名」
        r'『([^』]{2,20})』',  # 『商品名』
        r'#(\S{2,15})',        # #ハッシュタグ
    ]:
        for match in re.findall(pattern, text):
            if match not in product_keywords:
                product_keywords[match] = {"count": 0, "total_likes": 0, "category": t.get("category", "?")}
            product_keywords[match]["count"] += 1
            product_keywords[match]["total_likes"] += t["metrics"]["likes"]

top_products = sorted(product_keywords.items(), key=lambda x: x[1]["total_likes"], reverse=True)[:10]
for name, d in top_products[:5]:
    insights.append({
        "priority": "MEDIUM",
        "category": f"話題のキーワード（{d['category']}）",
        "insight": f"「{name}」: {d['count']}回言及, 合計{compact(d['total_likes'])}いいね"
    })

# 5. Content format analysis
for cat, tweets in all_data.items():
    if len(tweets) < 3: continue
    media_tweets = [t for t in tweets if t.get("media") and len(t["media"]) > 0]
    text_tweets = [t for t in tweets if not t.get("media") or len(t["media"]) == 0]
    if media_tweets and text_tweets:
        avg_m = sum(t["metrics"]["likes"] for t in media_tweets) / len(media_tweets)
        avg_t = sum(t["metrics"]["likes"] for t in text_tweets) / len(text_tweets)
        winner = "写真/動画付き" if avg_m > avg_t else "テキスト"
        insights.append({
            "priority": "MEDIUM",
            "category": f"形式（{cat}）",
            "insight": f"メディア付き: 平均{compact(avg_m)}いいね({len(media_tweets)}件) / テキスト: 平均{compact(avg_t)}いいね({len(text_tweets)}件) → {winner}が有効"
        })

# 6. Actionable summary
insights.append({
    "priority": "HIGH",
    "category": "アクション提案",
    "insight": "Z世代にリーチするなら: (1)フード系は写真1枚+短文で爆発力がある (2)推し活グッズは即完売情報が拡散される (3)コスメはBefore/After+体験談が刺さる (4)ガジェットはQOL改善のリスト形式が保存される"
})

row = 3
ws_ins.append(["優先度", "カテゴリ", "インサイト"])
style_header(ws_ins, row, 3)

for ins in insights:
    row += 1
    ws_ins.cell(row=row, column=1, value=ins["priority"])
    ws_ins.cell(row=row, column=2, value=ins["category"])
    ws_ins.cell(row=row, column=3, value=ins["insight"])
    ws_ins.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    if ins["priority"] == "HIGH":
        for c in range(1, 4):
            ws_ins.cell(row=row, column=c).fill = INSIGHT_FILL
            ws_ins.cell(row=row, column=c).font = BOLD

ws_ins.column_dimensions["A"].width = 12
ws_ins.column_dimensions["B"].width = 30
ws_ins.column_dimensions["C"].width = 90
ws_ins.freeze_panes = "A4"

# Move insights sheet to front
wb.move_sheet("戦略的インサイト", offset=-7)

# --- Save ---
reports_dir = Path("/Users/naoterumaker2/0_AI/x_skill/reports") / datetime.now().strftime("%Y-%m-%d")
reports_dir.mkdir(parents=True, exist_ok=True)
out = reports_dir / f"genz-trending-items-{datetime.now().strftime('%Y%m%d')}.xlsx"
wb.save(str(out))
print(f"\nSaved: {out}", file=sys.stderr)
print(f"Sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
print(f"Total: {len(all_tweets)} tweets, {len(insights)} insights", file=sys.stderr)
