#!/usr/bin/env python3
"""
Claude Code & OpenClaw X/Twitter Full Research Report
- Merges JA/EN/OpenClaw search results
- Strategic insights with actionable recommendations
- Exports comprehensive xlsx to x_skill directory
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

# --- Styling ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
INSIGHT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
BOLD = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
NUM_FMT = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_width(ws, min_w=8, max_w=60):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)

def compact(n):
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000: return f"{n/1_000:.1f}K"
    return str(int(n))

# --- Load Data ---
print("Loading data...", file=sys.stderr)
ja_tweets = json.loads(Path("/tmp/cc-ja.json").read_text())
en_tweets = json.loads(Path("/tmp/cc-en.json").read_text())
oc_tweets = json.loads(Path("/tmp/oc.json").read_text())

print(f"Claude Code JA: {len(ja_tweets)} tweets", file=sys.stderr)
print(f"Claude Code EN: {len(en_tweets)} tweets", file=sys.stderr)
print(f"OpenClaw: {len(oc_tweets)} tweets", file=sys.stderr)

# Deduplicate
seen = set()
def dedupe(tweets):
    result = []
    for t in tweets:
        if t["id"] not in seen:
            seen.add(t["id"])
            result.append(t)
    return result

ja_tweets = dedupe(ja_tweets)
en_tweets = dedupe(en_tweets)
oc_tweets = dedupe(oc_tweets)

all_cc = ja_tweets + en_tweets
all_tweets = all_cc + oc_tweets

print(f"After dedupe - JA: {len(ja_tweets)}, EN: {len(en_tweets)}, OC: {len(oc_tweets)}", file=sys.stderr)

# --- Analysis ---
def stats(tweets):
    if not tweets: return {}
    likes = [t["metrics"]["likes"] for t in tweets]
    imps = [t["metrics"].get("impressions", 0) for t in tweets]
    return {
        "count": len(tweets),
        "total_likes": sum(likes),
        "avg_likes": sum(likes)/len(likes),
        "median_likes": sorted(likes)[len(likes)//2],
        "max_likes": max(likes),
        "total_imps": sum(imps),
        "avg_imps": sum(imps)/len(imps) if imps else 0,
    }

ja_stats = stats(ja_tweets)
en_stats = stats(en_tweets)
oc_stats = stats(oc_tweets)

# Post type labels (used in insights + sheets)
POST_TYPE_LABELS = {
    "quote": "引用ポスト",
    "x_article": "X記事",
    "article_link": "記事リンク",
    "media": "メディア",
    "text": "テキスト",
}

# --- Strategic Insights ---
def generate_insights():
    insights = []

    # 1. JA vs EN comparison
    if ja_stats.get("count") and en_stats.get("count"):
        insights.append({
            "priority": "HIGH",
            "category": "言語別エンゲージメント",
            "insight": f"日本語: {ja_stats['count']}件, 平均{compact(ja_stats['avg_likes'])}いいね / 英語: {en_stats['count']}件, 平均{compact(en_stats['avg_likes'])}いいね。{'日本語圏の方が平均エンゲージメントが高い' if ja_stats['avg_likes'] > en_stats['avg_likes'] else '英語圏の方がリーチ・エンゲージメントが広い'}。"
        })

    # 2. Content type (media vs text)
    for label, tweets in [("Claude Code全体", all_cc), ("日本語", ja_tweets), ("英語", en_tweets)]:
        media = [t for t in tweets if t.get("media") and len(t["media"]) > 0]
        text_only = [t for t in tweets if not t.get("media") or len(t["media"]) == 0]
        if media and text_only:
            avg_m = sum(t["metrics"]["likes"] for t in media) / len(media)
            avg_t = sum(t["metrics"]["likes"] for t in text_only) / len(text_only)
            winner = "メディア付き" if avg_m > avg_t else "テキストのみ"
            insights.append({
                "priority": "HIGH",
                "category": f"コンテンツ形式（{label}）",
                "insight": f"メディア付き: {len(media)}件, 平均{compact(avg_m)}いいね / テキストのみ: {len(text_only)}件, 平均{compact(avg_t)}いいね → {winner}が強い。"
            })

    # 3. Theme analysis
    themes = {
        "生産性革命（○○やらせたら終わった系）": ["確定申告", "税理士", "全部終わった", "saved hours", "built entire", "shipped", "浮いた", "productivity"],
        "ユーモア・ネタ・ミーム": ["飛びました", "bubble", "meme", "average", "sleeping on", "but for", "放置", "lmao", "imagine"],
        "How-to / Tips / ガイド": ["tips", "how to", "guide", "master", "best practice", "CLAUDE.md", "workflow", "方法", "完全ガイド"],
        "新機能・公式アナウンス": ["announcing", "agent team", "opus 4.6", "hackathon", "update", "new feature", "エージェントチーム", "launch"],
        "衝撃データ・統計": ["4%", "commits", "10,000", "10000", "stars", "million", "never", "二度と", "90%", "inflection"],
        "ツール・エコシステム連携": ["figma", "chrome", "extension", "dexter", "openclaw", "vibecode", "vscode", "browser"],
        "感情・哲学的考察": ["weird time", "sadness", "wonder", "crazy", "profound", "never write code", "もう書かない"],
        "無料・コスト系": ["free", "無料", "$50", "credit", "no api cost", "ローカル"],
    }
    theme_results = []
    for theme, keywords in themes.items():
        matching = [t for t in all_cc if any(kw.lower() in t["text"].lower() for kw in keywords)]
        if matching:
            avg_l = sum(t["metrics"]["likes"] for t in matching) / len(matching)
            top = max(matching, key=lambda x: x["metrics"]["likes"])
            theme_results.append((theme, len(matching), avg_l, top))

    theme_results.sort(key=lambda x: x[2], reverse=True)
    for theme, count, avg_l, top in theme_results[:5]:
        insights.append({
            "priority": "HIGH" if avg_l > 3000 else "MEDIUM",
            "category": "バズるテーマ",
            "insight": f"「{theme}」: {count}件, 平均{compact(avg_l)}いいね。代表: @{top['username']}（{compact(top['metrics']['likes'])}いいね）"
        })

    # 4. Top accounts to emulate (JA)
    user_map = {}
    for t in ja_tweets:
        u = t["username"]
        if u not in user_map:
            user_map[u] = {"name": t.get("name","?"), "followers": t.get("author_followers",0), "tweets": [], "total_likes": 0}
        user_map[u]["tweets"].append(t)
        user_map[u]["total_likes"] += t["metrics"]["likes"]
    top_ja = sorted(user_map.items(), key=lambda x: x[1]["total_likes"], reverse=True)[:5]
    for u, d in top_ja:
        avg = d["total_likes"] / len(d["tweets"])
        insights.append({
            "priority": "HIGH",
            "category": "真似すべき日本語アカウント",
            "insight": f"@{u}（{d['name']}）: フォロワー{compact(d['followers'])}、{len(d['tweets'])}件で合計{compact(d['total_likes'])}いいね（平均{compact(avg)}）"
        })

    # 5. Top accounts to emulate (EN)
    user_map = {}
    for t in en_tweets:
        u = t["username"]
        if u not in user_map:
            user_map[u] = {"name": t.get("name","?"), "followers": t.get("author_followers",0), "tweets": [], "total_likes": 0}
        user_map[u]["tweets"].append(t)
        user_map[u]["total_likes"] += t["metrics"]["likes"]
    top_en = sorted(user_map.items(), key=lambda x: x[1]["total_likes"], reverse=True)[:5]
    for u, d in top_en:
        avg = d["total_likes"] / len(d["tweets"])
        insights.append({
            "priority": "HIGH",
            "category": "真似すべき英語アカウント",
            "insight": f"@{u}（{d['name']}）: フォロワー{compact(d['followers'])}、{len(d['tweets'])}件で合計{compact(d['total_likes'])}いいね（平均{compact(avg)}）"
        })

    # 6. Posting time
    hours = {}
    for t in all_cc:
        try:
            h = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00")).hour
            if h not in hours: hours[h] = []
            hours[h].append(t["metrics"]["likes"])
        except: pass
    if hours:
        best = sorted(hours.items(), key=lambda x: sum(x[1])/len(x[1]), reverse=True)[:3]
        h_str = ", ".join(f"UTC {h}時（JST {(h+9)%24}時）" for h, _ in best)
        insights.append({
            "priority": "MEDIUM",
            "category": "投稿時間帯",
            "insight": f"エンゲージメントが最も高い時間帯: {h_str}"
        })

    # 7. Post length
    short = [t for t in all_cc if len(t["text"]) < 100]
    mid = [t for t in all_cc if 100 <= len(t["text"]) < 280]
    long = [t for t in all_cc if len(t["text"]) >= 280]
    length_data = []
    if short: length_data.append(("短文(<100字)", sum(t["metrics"]["likes"] for t in short)/len(short), len(short)))
    if mid: length_data.append(("中文(100-280字)", sum(t["metrics"]["likes"] for t in mid)/len(mid), len(mid)))
    if long: length_data.append(("長文(280字+)", sum(t["metrics"]["likes"] for t in long)/len(long), len(long)))
    if length_data:
        best = max(length_data, key=lambda x: x[1])
        insights.append({
            "priority": "MEDIUM",
            "category": "文字数",
            "insight": f"最もバズる文字数: {best[0]}（平均{compact(best[1])}いいね、{best[2]}件）"
        })

    # 8. Viral pattern
    top10 = sorted(all_cc, key=lambda t: t["metrics"]["likes"], reverse=True)[:10]
    patterns = []
    for t in top10:
        text = t["text"]
        if len(text) < 80: patterns.append("短文インパクト")
        if "?" in text or "？" in text: patterns.append("疑問形")
        if any(w in text.lower() for w in ["never", "二度と", "もう", "all", "every", "全部"]): patterns.append("断言型")
        if any(w in text.lower() for w in ["just", "今", "breaking", "announced"]): patterns.append("速報型")
        if any(w in text.lower() for w in ["how to", "tips", "guide", "master"]): patterns.append("ハウツー型")
        if any(w in text.lower() for w in ["this is", "これが", "literally"]): patterns.append("「これが○○」型")
    pc = Counter(patterns)
    if pc:
        top_p = pc.most_common(3)
        insights.append({
            "priority": "HIGH",
            "category": "TOP10バズパターン",
            "insight": f"上位10投稿の共通パターン: {', '.join(f'{p}({c}件)' for p,c in top_p)}。この型を意識して投稿すべき。"
        })

    # 9. Post type engagement analysis
    for label, tweets in [("Claude Code全体", all_cc), ("日本語", ja_tweets), ("英語", en_tweets)]:
        type_stats = {}
        for t in tweets:
            pt = t.get("post_type", "text")
            if pt not in type_stats:
                type_stats[pt] = {"likes": [], "count": 0}
            type_stats[pt]["likes"].append(t["metrics"]["likes"])
            type_stats[pt]["count"] += 1
        if len(type_stats) >= 2:
            ranked = sorted(type_stats.items(), key=lambda x: sum(x[1]["likes"])/len(x[1]["likes"]), reverse=True)
            parts = []
            for pt, st in ranked:
                avg = sum(st["likes"]) / len(st["likes"])
                parts.append(f"{POST_TYPE_LABELS.get(pt, pt)}: {st['count']}件, 平均{compact(avg)}いいね")
            insights.append({
                "priority": "HIGH",
                "category": f"投稿タイプ別エンゲージメント（{label}）",
                "insight": " / ".join(parts)
            })

    # 10. X Article specific insight
    x_articles = [t for t in all_cc if t.get("post_type") == "x_article"]
    non_articles = [t for t in all_cc if t.get("post_type") != "x_article"]
    if x_articles:
        avg_xa = sum(t["metrics"]["likes"] for t in x_articles) / len(x_articles)
        avg_other = sum(t["metrics"]["likes"] for t in non_articles) / len(non_articles) if non_articles else 0
        ratio = avg_xa / avg_other if avg_other > 0 else 0
        insights.append({
            "priority": "HIGH",
            "category": "X記事（長文投稿）トレンド",
            "insight": f"X記事: {len(x_articles)}件, 平均{compact(avg_xa)}いいね（他形式比{ratio:.1f}倍）。{'X記事は通常投稿よりエンゲージメントが高い。長文で深い解説が刺さる傾向。' if ratio > 1 else 'X記事は他形式と同等。内容の質が重要。'}"
        })

    # 11. OpenClaw insights
    if oc_stats.get("count"):
        insights.append({
            "priority": "MEDIUM",
            "category": "OpenClaw動向",
            "insight": f"OpenClaw関連: {oc_stats['count']}件, 平均{compact(oc_stats['avg_likes'])}いいね, 最大{compact(oc_stats['max_likes'])}いいね。Claude Codeとの組み合わせ言及が多い。"
        })

    return insights

insights = generate_insights()

# --- Build Workbook ---
wb = Workbook()
REPORTS_DIR = Path("/Users/naoterumaker2/0_AI/x_skill/reports") / datetime.now().strftime("%Y-%m-%d")
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Sheet 1: Strategic Insights
ws = wb.active
ws.title = "戦略的インサイト"
ws.cell(row=1, column=1, value="Claude Code & OpenClaw バズ分析 - 戦略的インサイト").font = TITLE_FONT
ws.merge_cells("A1:C1")
ws.cell(row=2, column=1, value=f"分析日: {datetime.now().strftime('%Y-%m-%d')} | JA: {len(ja_tweets)}件 | EN: {len(en_tweets)}件 | OpenClaw: {len(oc_tweets)}件")
ws.merge_cells("A2:C2")

row = 4
ws.append(["優先度", "カテゴリ", "インサイト（アクション）"])
style_header(ws, row, 3)

for ins in insights:
    row += 1
    ws.cell(row=row, column=1, value=ins["priority"])
    ws.cell(row=row, column=2, value=ins["category"])
    ws.cell(row=row, column=3, value=ins["insight"])
    ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    if ins["priority"] == "HIGH":
        for c in range(1,4):
            ws.cell(row=row, column=c).fill = INSIGHT_FILL
            ws.cell(row=row, column=c).font = BOLD

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 90
ws.freeze_panes = "A5"

POST_TYPE_FILLS = {
    "quote": PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
    "x_article": PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
    "article_link": ORANGE_FILL,
    "media": GREEN_FILL,
    "text": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
}

# Sheet helper
VIRAL_FILL = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")

def write_tweet_sheet(ws, tweets, title):
    # アカウントURLはUsername横、ポストURLはテキスト横、効率指標はエンゲージメント横
    headers = ["No", "投稿タイプ", "Username", "アカウントURL", "Name", "Followers",
               "テキスト", "ポストURL", "いいね", "RT", "リプ", "インプレッション", "ブックマーク",
               "バズ効率", "保存率",
               "メディア種別", "投稿日時"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for i, t in enumerate(sorted(tweets, key=lambda x: x["metrics"]["likes"], reverse=True), 1):
        m = t["metrics"]
        media_types = ", ".join(md.get("type","") for md in t.get("media", []))
        post_type = t.get("post_type", "text")
        followers = t.get("author_followers", 0) or 1
        likes = m["likes"]
        bookmarks = m.get("bookmarks", 0)
        buzz_eff = likes / followers
        save_rate = bookmarks / likes if likes > 0 else 0
        ws.append([
            i,
            POST_TYPE_LABELS.get(post_type, post_type),
            f"@{t['username']}",
            t.get("account_url", f"https://x.com/{t['username']}"),
            t.get("name","?"),
            followers,
            t["text"][:300],
            t.get("tweet_url",""),
            likes, m["retweets"], m["replies"],
            m.get("impressions",0), bookmarks,
            round(buzz_eff, 2), f"{save_rate:.1%}",
            media_types, t.get("created_at",""),
        ])
        row_num = ws.max_row
        for col_idx in [6,9,10,11,12,13]:
            ws.cell(row=row_num, column=col_idx).number_format = NUM_FMT
        # Highlight viral efficiency >= 1.0
        if buzz_eff >= 1.0:
            ws.cell(row=row_num, column=14).fill = VIRAL_FILL
            ws.cell(row=row_num, column=14).font = BOLD
        # Color post type
        fill = POST_TYPE_FILLS.get(post_type)
        if fill:
            ws.cell(row=row_num, column=2).fill = fill
    auto_width(ws, max_w=50)
    ws.column_dimensions["D"].width = 30   # アカウントURL
    ws.column_dimensions["G"].width = 60   # テキスト
    ws.column_dimensions["H"].width = 45   # ポストURL
    ws.column_dimensions["N"].width = 12   # バズ効率
    ws.column_dimensions["O"].width = 10   # 保存率
    ws.freeze_panes = "A2"

# Sheet 2: JA tweets
ws_ja = wb.create_sheet("日本語バズ投稿")
write_tweet_sheet(ws_ja, ja_tweets, "日本語")

# Sheet 3: EN tweets
ws_en = wb.create_sheet("英語バズ投稿")
write_tweet_sheet(ws_en, en_tweets, "英語")

# Sheet 4: OpenClaw tweets
ws_oc = wb.create_sheet("OpenClaw")
write_tweet_sheet(ws_oc, oc_tweets, "OpenClaw")

# Sheet 5: Account Rankings
ws_rank = wb.create_sheet("アカウントランキング")
ws_rank.cell(row=1, column=1, value="バズアカウントTOP（真似すべき投稿スタイル）").font = TITLE_FONT
ws_rank.merge_cells("A1:J1")

row = 3
ws_rank.append(["言語", "Username", "Name", "Followers", "投稿数", "合計いいね", "平均いいね", "主な投稿タイプ", "トップ投稿", "アカウントURL"])
style_header(ws_rank, row, 10)

user_all = {}
for lang, tweets in [("JA", ja_tweets), ("EN", en_tweets), ("OC", oc_tweets)]:
    for t in tweets:
        key = (lang, t["username"])
        if key not in user_all:
            user_all[key] = {"lang": lang, "name": t.get("name","?"), "followers": t.get("author_followers",0), "tweets": [], "total_likes": 0}
        user_all[key]["tweets"].append(t)
        user_all[key]["total_likes"] += t["metrics"]["likes"]

sorted_users = sorted(user_all.items(), key=lambda x: x[1]["total_likes"], reverse=True)
for (lang, username), d in sorted_users[:40]:
    row += 1
    avg = d["total_likes"] / len(d["tweets"])
    top = max(d["tweets"], key=lambda x: x["metrics"]["likes"])
    # Determine dominant post type for this account
    type_counts = Counter(t.get("post_type", "text") for t in d["tweets"])
    dominant_types = ", ".join(POST_TYPE_LABELS.get(pt, pt) for pt, _ in type_counts.most_common(2))
    ws_rank.cell(row=row, column=1, value=lang)
    ws_rank.cell(row=row, column=2, value=f"@{username}")
    ws_rank.cell(row=row, column=3, value=d["name"])
    ws_rank.cell(row=row, column=4, value=d["followers"])
    ws_rank.cell(row=row, column=5, value=len(d["tweets"]))
    ws_rank.cell(row=row, column=6, value=d["total_likes"])
    ws_rank.cell(row=row, column=7, value=int(avg))
    ws_rank.cell(row=row, column=8, value=dominant_types)
    ws_rank.cell(row=row, column=9, value=f"{top['text'][:120]}")
    ws_rank.cell(row=row, column=10, value=f"https://x.com/{username}")
    for c in [4,6,7]: ws_rank.cell(row=row, column=c).number_format = NUM_FMT
    ws_rank.cell(row=row, column=9).alignment = Alignment(wrap_text=True)
    if lang == "JA": ws_rank.cell(row=row, column=1).fill = GREEN_FILL
    elif lang == "EN": ws_rank.cell(row=row, column=1).fill = ACCENT_FILL
    else: ws_rank.cell(row=row, column=1).fill = ORANGE_FILL

auto_width(ws_rank, max_w=50)
ws_rank.column_dimensions["I"].width = 80
ws_rank.column_dimensions["J"].width = 30
ws_rank.freeze_panes = "A4"

# Sheet 6: Theme Analysis
ws_th = wb.create_sheet("テーマ分析")
ws_th.cell(row=1, column=1, value="テーマ別バズ傾向分析").font = TITLE_FONT
ws_th.merge_cells("A1:E1")

themes = {
    "生産性革命（○○やらせたら終わった系）": ["確定申告", "税理士", "全部終わった", "saved hours", "built entire", "shipped", "浮いた"],
    "ユーモア・ネタ・ミーム": ["飛びました", "bubble", "meme", "average", "sleeping on", "but for", "放置"],
    "How-to / Tips / ガイド": ["tips", "how to", "guide", "master", "best practice", "CLAUDE.md", "workflow", "完全ガイド"],
    "新機能・公式アナウンス": ["announcing", "agent team", "opus 4.6", "hackathon", "エージェントチーム", "launch"],
    "衝撃データ・統計": ["4%", "commits", "10,000", "stars", "million", "never", "二度と", "90%", "inflection"],
    "ツール・エコシステム連携": ["figma", "chrome", "extension", "dexter", "openclaw", "vibecode", "vscode"],
    "感情・哲学的考察": ["weird time", "sadness", "wonder", "never write code", "もう書かない", "profound"],
    "無料・コスト系": ["free", "無料", "$50", "credit", "no api cost", "ローカル", "open source"],
}

row = 3
ws_th.append(["テーマ", "ヒット数", "平均いいね", "最大いいね", "代表ポスト"])
style_header(ws_th, row, 5)

for theme, kws in themes.items():
    matching = [t for t in all_cc if any(kw.lower() in t["text"].lower() for kw in kws)]
    if matching:
        avg_l = sum(t["metrics"]["likes"] for t in matching) / len(matching)
        top = max(matching, key=lambda x: x["metrics"]["likes"])
        row += 1
        ws_th.cell(row=row, column=1, value=theme)
        ws_th.cell(row=row, column=2, value=len(matching))
        ws_th.cell(row=row, column=3, value=int(avg_l))
        ws_th.cell(row=row, column=4, value=top["metrics"]["likes"])
        ws_th.cell(row=row, column=5, value=f"@{top['username']}: {top['text'][:120]}")
        for c in [3,4]: ws_th.cell(row=row, column=c).number_format = NUM_FMT
        ws_th.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

auto_width(ws_th, max_w=50)
ws_th.column_dimensions["E"].width = 80

# Sheet 7: タイプ×言語クロス集計
ws_cross = wb.create_sheet("タイプ×言語クロス集計")
ws_cross.cell(row=1, column=1, value="投稿タイプ × 言語別エンゲージメント比較").font = TITLE_FONT
ws_cross.merge_cells("A1:H1")

# Build cross table data
cross_data = {}
for lang_label, tweets in [("日本語", ja_tweets), ("英語", en_tweets), ("OpenClaw", oc_tweets)]:
    for t in tweets:
        pt = t.get("post_type", "text")
        key = (POST_TYPE_LABELS.get(pt, pt), lang_label)
        if key not in cross_data:
            cross_data[key] = {"likes": [], "bookmarks": [], "count": 0}
        cross_data[key]["likes"].append(t["metrics"]["likes"])
        cross_data[key]["bookmarks"].append(t["metrics"].get("bookmarks", 0))
        cross_data[key]["count"] += 1

pt_order = ["X記事", "メディア", "引用ポスト", "記事リンク", "テキスト"]
lang_order = ["日本語", "英語", "OpenClaw"]

row = 3
# Header row 1: metric labels
ws_cross.cell(row=row, column=1, value="投稿タイプ")
col = 2
for lang in lang_order:
    ws_cross.cell(row=row, column=col, value=f"{lang} 件数")
    ws_cross.cell(row=row, column=col+1, value=f"{lang} 平均いいね")
    ws_cross.cell(row=row, column=col+2, value=f"{lang} 平均保存率")
    col += 3
style_header(ws_cross, row, col - 1)

for pt_label in pt_order:
    row += 1
    ws_cross.cell(row=row, column=1, value=pt_label)
    ws_cross.cell(row=row, column=1).font = BOLD
    col = 2
    best_avg = 0
    best_col = -1
    row_avgs = []
    for lang in lang_order:
        d = cross_data.get((pt_label, lang))
        if d and d["count"] > 0:
            avg_l = sum(d["likes"]) / len(d["likes"])
            total_l = sum(d["likes"])
            total_b = sum(d["bookmarks"])
            save_rate = total_b / total_l if total_l > 0 else 0
            ws_cross.cell(row=row, column=col, value=d["count"])
            ws_cross.cell(row=row, column=col+1, value=int(avg_l))
            ws_cross.cell(row=row, column=col+1).number_format = NUM_FMT
            ws_cross.cell(row=row, column=col+2, value=f"{save_rate:.1%}")
            row_avgs.append((col+1, avg_l))
        else:
            ws_cross.cell(row=row, column=col, value="-")
            ws_cross.cell(row=row, column=col+1, value="-")
            ws_cross.cell(row=row, column=col+2, value="-")
        col += 3
    # Highlight best performing language for this type
    if row_avgs:
        best_col_idx = max(row_avgs, key=lambda x: x[1])[0]
        ws_cross.cell(row=row, column=best_col_idx).fill = INSIGHT_FILL
        ws_cross.cell(row=row, column=best_col_idx).font = BOLD

auto_width(ws_cross, min_w=10, max_w=20)
ws_cross.column_dimensions["A"].width = 16
ws_cross.freeze_panes = "B4"

# Sheet 8: バズ公式テンプレート
ws_tmpl = wb.create_sheet("バズ公式テンプレート")
ws_tmpl.cell(row=1, column=1, value="バズ公式テンプレート - TOP投稿から抽出した再現可能な型").font = TITLE_FONT
ws_tmpl.merge_cells("A1:F1")

# Define templates extracted from actual top posts
TEMPLATES = [
    {
        "formula": "○○に○○やらせたら全部終わった",
        "lang": "JA",
        "type": "生産性革命",
        "example_kw": ["やらせたら", "全部終わった", "浮いた", "不要になった", "秒で"],
        "tip": "身近なコスト（税理士、デザイナー等）を具体的に。短文+驚きが鍵。",
    },
    {
        "formula": "Claude Code but for ○○",
        "lang": "EN",
        "type": "ミーム・ネタ",
        "example_kw": ["but for", "claude code but"],
        "tip": "意外な領域（hacking, cooking等）+ 動画/GIFデモ必須。",
    },
    {
        "formula": "○○に○○を接続してみた + デモ動画",
        "lang": "Both",
        "type": "衝撃デモ",
        "example_kw": ["接続", "hooked up", "installed", "pointed it at"],
        "tip": "ハードウェア連携が強い。カメラ、スマホ、IoT等。動画必須。",
    },
    {
        "formula": "○○が○○した（擬人化・ネタ系）",
        "lang": "JA",
        "type": "ユーモア",
        "example_kw": ["飛びました", "放置", "「なにが？笑」"],
        "tip": "AIの予想外の行動をスクショ付きで。ツッコミ要素が重要。",
    },
    {
        "formula": "This is literally everything you need to ○○",
        "lang": "EN",
        "type": "ガイド・まとめ",
        "example_kw": ["everything you need", "master", "complete guide"],
        "tip": "画像1枚にまとめたチートシート形式。保存率が非常に高い。",
    },
    {
        "formula": "○○を無料で学べる○○が公開 + 箇条書き",
        "lang": "JA",
        "type": "情報まとめ",
        "example_kw": ["無料で学べる", "公開", "学習内容"],
        "tip": "学習系コンテンツ+箇条書きでブックマーク率UP。",
    },
    {
        "formula": "just talked with ○○ / never heard of ○○ / we are in a bubble",
        "lang": "EN",
        "type": "業界考察",
        "example_kw": ["bubble", "never heard of", "sleeping on", "inflection"],
        "tip": "テック業界の温度差を示す実体験。短文で断言が刺さる。",
    },
    {
        "formula": "Anthropicが○○を公開した（速報解説X記事）",
        "lang": "JA",
        "type": "速報X記事",
        "example_kw": ["公開した", "発表した", "リリース"],
        "tip": "公式発表を即座にX記事で日本語解説。先行者利益が大きい。",
    },
]

row = 3
ws_tmpl.append(["No", "バズ公式", "言語", "カテゴリ", "使い方のコツ", "キーワード例"])
style_header(ws_tmpl, row, 6)

for i, tmpl in enumerate(TEMPLATES, 1):
    row += 1
    ws_tmpl.cell(row=row, column=1, value=i)
    ws_tmpl.cell(row=row, column=2, value=tmpl["formula"])
    ws_tmpl.cell(row=row, column=2).font = BOLD
    ws_tmpl.cell(row=row, column=3, value=tmpl["lang"])
    ws_tmpl.cell(row=row, column=4, value=tmpl["type"])
    ws_tmpl.cell(row=row, column=5, value=tmpl["tip"])
    ws_tmpl.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
    ws_tmpl.cell(row=row, column=6, value=", ".join(tmpl["example_kw"]))
    # Color by language
    if tmpl["lang"] == "JA":
        ws_tmpl.cell(row=row, column=3).fill = GREEN_FILL
    elif tmpl["lang"] == "EN":
        ws_tmpl.cell(row=row, column=3).fill = ACCENT_FILL
    else:
        ws_tmpl.cell(row=row, column=3).fill = ORANGE_FILL
    # Match actual posts to this template
    matching = [t for t in all_tweets if any(kw.lower() in t["text"].lower() for kw in tmpl["example_kw"])]
    if matching:
        avg_l = sum(t["metrics"]["likes"] for t in matching) / len(matching)
        best = max(matching, key=lambda x: x["metrics"]["likes"])
        ws_tmpl.cell(row=row, column=2).value = f'{tmpl["formula"]}  [{len(matching)}件, 平均{compact(avg_l)}いいね]'

# Add actual example rows under each template
row += 2
ws_tmpl.cell(row=row, column=1, value="実績TOP: バズ効率が高い投稿").font = TITLE_FONT
ws_tmpl.merge_cells(f"A{row}:F{row}")

row += 1
ws_tmpl.append(["順位", "Username", "テキスト", "いいね", "フォロワー", "バズ効率"])
style_header(ws_tmpl, row, 6)

# Sort by buzz efficiency
eff_sorted = sorted(all_tweets, key=lambda t: t["metrics"]["likes"] / max(t.get("author_followers", 1), 1), reverse=True)
for i, t in enumerate(eff_sorted[:15], 1):
    row += 1
    followers = t.get("author_followers", 1) or 1
    eff = t["metrics"]["likes"] / followers
    ws_tmpl.cell(row=row, column=1, value=i)
    ws_tmpl.cell(row=row, column=2, value=f"@{t['username']}")
    ws_tmpl.cell(row=row, column=3, value=t["text"][:150])
    ws_tmpl.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    ws_tmpl.cell(row=row, column=4, value=t["metrics"]["likes"])
    ws_tmpl.cell(row=row, column=4).number_format = NUM_FMT
    ws_tmpl.cell(row=row, column=5, value=followers)
    ws_tmpl.cell(row=row, column=5).number_format = NUM_FMT
    ws_tmpl.cell(row=row, column=6, value=round(eff, 2))
    if eff >= 1.0:
        ws_tmpl.cell(row=row, column=6).fill = VIRAL_FILL
        ws_tmpl.cell(row=row, column=6).font = BOLD

ws_tmpl.column_dimensions["A"].width = 8
ws_tmpl.column_dimensions["B"].width = 45
ws_tmpl.column_dimensions["C"].width = 60
ws_tmpl.column_dimensions["D"].width = 12
ws_tmpl.column_dimensions["E"].width = 30
ws_tmpl.column_dimensions["F"].width = 20

# --- Save ---
out = REPORTS_DIR / f"claude-code-openclaw-buzz-{datetime.now().strftime('%Y%m%d')}.xlsx"
wb.save(str(out))
print(f"\nSaved: {out}", file=sys.stderr)
print(f"Sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
print(f"\nSummary:", file=sys.stderr)
print(f"  JA: {len(ja_tweets)} tweets (100+ likes)", file=sys.stderr)
print(f"  EN: {len(en_tweets)} tweets (100+ likes)", file=sys.stderr)
print(f"  OpenClaw: {len(oc_tweets)} tweets (50+ likes)", file=sys.stderr)
print(f"  Insights: {len(insights)} items", file=sys.stderr)
