#!/usr/bin/env python3
"""
Generate formatted xlsx from JSON export data.
4 sheets: Tweets, Engagement, Influencers, Keywords & Sentiment.
Requires: openpyxl >= 3.1.5
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def style_header(ws, row, cols):
    """Apply header styling."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="999999"),
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def auto_width(ws, min_width=8, max_width=50):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def write_tweets_sheet(wb, tweets):
    ws = wb.active
    ws.title = "Tweets"

    headers = [
        "No", "Username", "Name", "Followers", "Text", "Likes", "RT",
        "Replies", "Impressions", "Bookmarks", "Eng.Rate", "Media Type",
        "Media URL", "Link Title", "Posted At", "URL"
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    num_fmt = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    pct_fmt = '0.00%'

    for i, t in enumerate(tweets, 1):
        metrics = t.get("metrics", {})
        imp = metrics.get("impressions", 0)
        eng = metrics.get("likes", 0) + metrics.get("retweets", 0) + metrics.get("replies", 0)
        eng_rate = eng / imp if imp > 0 else 0

        media = t.get("media", [])
        media_type = media[0]["type"] if media else ""
        media_url = media[0].get("url", media[0].get("preview_url", "")) if media else ""

        url_meta = t.get("url_meta", [])
        link_title = url_meta[0].get("title", "") if url_meta else ""

        row = [
            i,
            f"@{t.get('username', '?')}",
            t.get("name", "?"),
            t.get("author_followers", 0),
            t.get("text", ""),
            metrics.get("likes", 0),
            metrics.get("retweets", 0),
            metrics.get("replies", 0),
            metrics.get("impressions", 0),
            metrics.get("bookmarks", 0),
            eng_rate,
            media_type,
            media_url,
            link_title,
            t.get("created_at", ""),
            t.get("tweet_url", ""),
        ]
        ws.append(row)

    # Format columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.number_format = pct_fmt

    for col_idx in [4, 6, 7, 8, 9, 10]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = num_fmt

    # Summary row
    if tweets:
        last_row = len(tweets) + 2
        ws.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)
        for col_idx, label in [(6, "likes"), (7, "retweets"), (8, "replies"), (9, "impressions"), (10, "bookmarks")]:
            total = sum(t.get("metrics", {}).get(label, 0) for t in tweets)
            cell = ws.cell(row=last_row, column=col_idx, value=total)
            cell.font = Font(bold=True)
            cell.number_format = num_fmt

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_engagement_sheet(wb, engagement):
    ws = wb.create_sheet("Engagement")

    # Stats summary
    headers = ["Metric", "Average", "Median", "Max", "Total"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for metric_name in ["likes", "impressions", "retweets", "replies"]:
        m = engagement.get(metric_name, {})
        ws.append([
            metric_name.capitalize(),
            round(m.get("avg", 0), 1),
            round(m.get("median", 0), 1),
            m.get("max", 0),
            m.get("total", 0),
        ])

    eng_rate = engagement.get("engagementRate", 0)
    ws.append([])
    ws.append(["Engagement Rate", f"{eng_rate * 100:.2f}%"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    # Content type breakdown
    ws.append([])
    ws.append(["Content Type", "Count", "Avg Likes", "Avg Impressions"])
    style_header(ws, ws.max_row, 4)

    ct = engagement.get("byContentType", {})
    for label, key in [("Text Only", "textOnly"), ("With Media", "withMedia"), ("With Links", "withLinks")]:
        data = ct.get(key, {})
        ws.append([
            label,
            data.get("count", 0),
            round(data.get("avgLikes", 0), 1),
            round(data.get("avgImpressions", 0), 1),
        ])

    # Top hours
    by_hour = engagement.get("byHour", {})
    if by_hour:
        ws.append([])
        ws.append(["Hour (UTC)", "Tweets", "Avg Likes", "Avg Impressions"])
        style_header(ws, ws.max_row, 4)
        for hour in sorted(by_hour.keys(), key=lambda h: by_hour[h].get("avgLikes", 0), reverse=True):
            data = by_hour[hour]
            ws.append([
                f"{int(hour):02d}:00",
                data.get("count", 0),
                round(data.get("avgLikes", 0), 1),
                round(data.get("avgImpressions", 0), 1),
            ])

    auto_width(ws)


def write_influencers_sheet(wb, influencers):
    ws = wb.create_sheet("Influencers")

    headers = ["Username", "Name", "Followers", "Following", "Tweets", "Total Likes",
               "Total Impressions", "Total RT", "Eng. Rate", "Category"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    cat_fills = {
        "high_follower": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "emerging_voice": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "regular": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    }

    for inf in influencers:
        row_data = [
            f"@{inf.get('username', '?')}",
            inf.get("name", "?"),
            inf.get("followers", 0),
            inf.get("following", 0),
            inf.get("tweetCount", 0),
            inf.get("totalLikes", 0),
            inf.get("totalImpressions", 0),
            inf.get("totalRetweets", 0),
            inf.get("avgEngagementRate", 0),
            inf.get("category", "regular"),
        ]
        ws.append(row_data)
        row_num = ws.max_row
        cat = inf.get("category", "regular")
        fill = cat_fills.get(cat, cat_fills["regular"])
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill
        ws.cell(row=row_num, column=9).number_format = '0.00%'

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_keywords_sheet(wb, keywords):
    ws = wb.create_sheet("Keywords & Sentiment")

    # Top keywords
    ws.append(["Keyword", "Count"])
    style_header(ws, 1, 2)
    for kw in keywords.get("topWords", [])[:30]:
        ws.append([kw["word"], kw["count"]])

    # Hashtags
    ws.append([])
    ws.append(["Hashtag", "Count"])
    style_header(ws, ws.max_row, 2)
    for ht in keywords.get("topHashtags", []):
        ws.append([f"#{ht['tag']}", ht["count"]])

    # URLs
    ws.append([])
    ws.append(["Shared URL", "Title", "Count"])
    style_header(ws, ws.max_row, 3)
    for u in keywords.get("topUrls", []):
        ws.append([u["url"], u["title"], u["count"]])

    # Sentiment
    ws.append([])
    sentiment = keywords.get("sentiment", {})
    ws.append(["Sentiment", "Count", "Keywords"])
    style_header(ws, ws.max_row, 3)
    ws.append(["Positive", sentiment.get("positive", 0),
               ", ".join(sentiment.get("positiveKeywords", [])[:10])])
    pos_row = ws.max_row
    ws.cell(row=pos_row, column=1).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    ws.append(["Negative", sentiment.get("negative", 0),
               ", ".join(sentiment.get("negativeKeywords", [])[:10])])
    neg_row = ws.max_row
    ws.cell(row=neg_row, column=1).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    ws.append(["Neutral", sentiment.get("neutral", 0), ""])

    auto_width(ws)


def main():
    if len(sys.argv) < 3:
        print("Usage: xlsx_export.py <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    with open(input_path, "r") as f:
        data = json.load(f)

    wb = Workbook()

    write_tweets_sheet(wb, data.get("tweets", []))
    write_engagement_sheet(wb, data.get("engagement", {}))
    write_influencers_sheet(wb, data.get("influencers", []))
    write_keywords_sheet(wb, data.get("keywords", {}))

    wb.save(str(output_path))
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
