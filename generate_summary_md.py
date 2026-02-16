#!/usr/bin/env python3
"""
X Research → Markdown + xlsx バズ分析レポート
JSON検索結果から「何が語られているか」を中心にmd + xlsxに出力。

Usage:
  python3 generate_summary_md.py --name "テーマ名" --files /tmp/a.json /tmp/b.json \
    --labels "ラベルA" "ラベルB" --queries 'query1' 'query2' \
    --titles /tmp/titles.json
"""

import json, sys, argparse, re
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    from openpyxl.utils import get_column_letter

def compact(n):
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000: return f"{n/1_000:.1f}K"
    return str(int(n))

POST_TYPE_LABELS = {
    "quote": "引用", "x_article": "X記事", "article_link": "記事リンク",
    "media": "メディア", "text": "テキスト",
}

# ============================================================
# 話題検出
# ============================================================

TOPIC_RULES = [
    ("LP/Web制作", ["lp", "ランディング", "figma", "web制作", "デザインデータ", "コーディング", "html", "css", "webサイト", "ホームページ"]),
    ("SEO/検索流入", ["seo", "検索", "オーガニック", "discover", "google", "インデックス", "被リンク", "ドメイン", "organic", "serp", "keyword", "rank"]),
    ("AI活用/テック", ["claude", "chatgpt", "gpt", "gemini", "cursor", "ai", "code", "プロンプト", "api", "llm", "opus", "notebooklm", "生成ai"]),
    ("コンテンツ制作", ["ライティング", "記事", "ブログ", "コンテンツ", "文章", "執筆", "原稿", "レビュー", "漫画", "kindle"]),
    ("AI副業/収益化", ["副業", "稼", "収益", "自動化", "放置", "不労", "月収", "借金", "脱サラ", "完全チート", "マネタイズ"]),
    ("ビジネス/起業", ["起業", "売上", "事業", "会社", "経営", "ベンチャー", "スタートアップ", "ceo", "ユニコーン", "企業価値"]),
    ("𝕏攻略/SNS", ["インプ", "アルゴリズム", "フォロワー", "バズ", "伸び", "note ", "𝕏", "x記事", "ポスト", "sns", "交流"]),
    ("広告/集客", ["広告", "集客", "cvr", "コンバージョン", "リスティング", "facebook広告", "instagram", "運用型"]),
    ("速報/ニュース", ["速報", "新機能", "リリース", "公開", "breaking", "ベータ", "発表", "コアアプデ"]),
]

def detect_topics(t):
    """テキストから話題を検出"""
    text = t.get("text", "").lower()
    if re.match(r'^https?://t\.co/\S+$', text.strip()):
        # X記事等でテキストなし → タイトルがあれば使う
        title = t.get("_title", "")
        if title:
            text = title.lower()
        else:
            return []
    # 短いキーワード（4文字以下の英字のみ）はワードバウンダリで検索
    SHORT_EN_RE = re.compile(r'^[a-z]{1,4}$')
    topics = []
    for topic, keywords in TOPIC_RULES:
        matched = False
        for kw in keywords:
            if SHORT_EN_RE.match(kw):
                if re.search(r'\b' + re.escape(kw) + r'\b', text):
                    matched = True
                    break
            elif kw in text:
                matched = True
                break
        if matched:
            topics.append(topic)
    return topics

# ============================================================
# X記事検出 & post_type修正
# ============================================================

def is_x_article_url(url):
    return bool(re.search(r'x\.com/(i/article|[^/]+/articles?)/', url))

def fix_post_type(t):
    urls = t.get("urls", [])
    if isinstance(urls, list) and any(is_x_article_url(u) for u in urls if isinstance(u, str)):
        t["post_type"] = "x_article"
    text = t.get("text", "").strip()
    if re.match(r'^https?://t\.co/\S+$', text):
        if t.get("post_type") != "x_article":
            for u in urls:
                if isinstance(u, str) and is_x_article_url(u):
                    t["post_type"] = "x_article"
                    break
    return t

def get_article_url(t):
    """X記事の実際のURLを取得（t.coではなく）"""
    for u in t.get("urls", []):
        if isinstance(u, str) and is_x_article_url(u):
            return u
    return None

def get_display_text(t, max_len=0):
    """投稿の表示テキストを生成"""
    text = t.get("text", "").strip()
    pt = t.get("post_type", "text")
    title = t.get("_title", "")

    if re.match(r'^https?://t\.co/\S+$', text):
        if title:
            return f"「{title}」"
        article_url = get_article_url(t)
        if pt == "x_article":
            if article_url:
                return f"[X記事] タイトル未取得 → {article_url}"
            return "[X記事] タイトル未取得"
        elif pt == "media":
            return "[メディア投稿] ※画像/動画はポストを参照"
        else:
            return "[リンク投稿] ※内容はポストを参照"

    if max_len and len(text) > max_len:
        return text[:max_len] + "…"
    return text

def tag_buzz_reason(t):
    tags = []
    text = t["text"].lower()
    raw_text = t["text"].strip()
    pt = t.get("post_type", "text")
    sr = t["metrics"].get("bookmarks", 0) / max(t["metrics"]["likes"], 1)
    is_url_only = bool(re.match(r'^https?://t\.co/\S+$', raw_text))

    if pt == "x_article": tags.append("X記事")
    elif t.get("media") and len(t.get("media", [])) > 0: tags.append("ビジュアル")
    if not is_url_only and len(raw_text) < 80: tags.append("短文一撃")

    if not is_url_only:
        if any(w in text for w in ["how to", "方法", "guide", "tips", "tutorial", "step", "コツ", "やり方", "入門", "まとめ"]): tags.append("ハウツー/まとめ")
        if any(w in text for w in ["$", "revenue", "earn", "稼", "profit", "made $", "income", "money", "年収", "売上"]): tags.append("収益系")
        if any(w in text for w in ["scared", "失敗", "lost", "怖", "mistake", "wrong", "regret", "倒産", "地獄"]): tags.append("体験談/リアル")
        if any(w in text for w in ["just", "今", "breaking", "公開", "shipped", "released", "announcing", "速報"]): tags.append("速報/リリース")
        if any(w in text for w in ["thread", "🧵", "ツリー"]): tags.append("スレッド")
        if "?" in raw_text or "？" in raw_text: tags.append("問いかけ")

    # タイトルがある場合もチェック
    title = t.get("_title", "").lower()
    if title:
        if any(w in title for w in ["方法", "まとめ", "入門", "コツ", "やり方"]): tags.append("ハウツー/まとめ")
        if any(w in title for w in ["年収", "稼", "売上", "金持ち"]): tags.append("収益系")
        if any(w in title for w in ["失敗", "倒産", "地獄"]): tags.append("体験談/リアル")

    if sr >= 1.0: tags.append("高保存率")
    return list(dict.fromkeys(tags)) if tags else ["—"]  # 重複除去


# ============================================================
# ノイズ自動検出
# ============================================================

_HAS_KANA = re.compile(r'[\u3040-\u309f\u30a0-\u30ff]')  # ひらがな or カタカナ

NOISE_PATTERNS = [
    ("KR", re.compile(r'[\uac00-\ud7af]')),       # 韓国語
    ("PT", re.compile(r'\b(desse|seria|dizem|estamos|vendo|nesses|também|porque|então)\b', re.I)),  # ポルトガル語
    ("ES", re.compile(r'\b(también|porque|entonces|después|nosotros|ustedes)\b', re.I)),  # スペイン語
    ("AR", re.compile(r'[\u0600-\u06ff]{5,}')),     # アラビア語
]

def detect_noise(t, target_langs=("ja", "en")):
    """非ターゲット言語のノイズを検出。lang_code or None"""
    text = t.get("text", "")
    # ひらがな/カタカナがあれば日本語 → ノイズではない
    if _HAS_KANA.search(text):
        return None
    for lang_code, pattern in NOISE_PATTERNS:
        if pattern.search(text):
            return lang_code
    return None


def load_and_dedupe(files, labels, title_map=None, exclude_ids=None, auto_noise=True):
    all_tweets = []
    noise_tweets = []
    seen = set(exclude_ids or set())
    per_label = {}
    for f, label in zip(files, labels):
        tweets = json.loads(Path(f).read_text())
        deduped = []
        for t in tweets:
            if t["id"] not in seen:
                seen.add(t["id"])
                t["_label"] = label
                fix_post_type(t)
                # タイトルマッピングを適用
                if title_map:
                    tid = t["id"]
                    url = t.get("tweet_url", "")
                    if tid in title_map:
                        t["_title"] = title_map[tid]
                    elif url in title_map:
                        t["_title"] = title_map[url]
                # ノイズ自動検出
                if auto_noise:
                    noise_lang = detect_noise(t)
                    if noise_lang:
                        noise_tweets.append((t, noise_lang))
                        continue
                deduped.append(t)
                all_tweets.append(t)
        per_label[label] = deduped

    if noise_tweets:
        print(f"[自動ノイズ除去] {len(noise_tweets)}件を除外:", file=sys.stderr)
        for t, lang in noise_tweets:
            print(f"  {lang} @{t.get('username','?')} ({t['metrics']['likes']}L): {t['text'][:50]}", file=sys.stderr)

    return all_tweets, per_label


# ============================================================
# 分析関数
# ============================================================

def analyze_topics(all_tweets):
    """話題マップを生成"""
    topic_tweets = defaultdict(list)
    for t in all_tweets:
        for topic in detect_topics(t):
            topic_tweets[topic].append(t)
    # いいね合計順
    return sorted(topic_tweets.items(), key=lambda x: sum(t["metrics"]["likes"] for t in x[1]), reverse=True)

def analyze_accounts(all_tweets):
    """アカウント別プロファイル"""
    by_user = {}
    for t in all_tweets:
        u = t.get("username", "?")
        if u not in by_user:
            by_user[u] = {
                "tweets": [], "followers": t.get("author_followers", 0),
                "account_url": t.get("account_url", ""),
            }
        by_user[u]["tweets"].append(t)

    profiles = []
    for username, data in by_user.items():
        tweets = data["tweets"]
        total_likes = sum(t["metrics"]["likes"] for t in tweets)
        total_bmarks = sum(t["metrics"].get("bookmarks", 0) for t in tweets)
        type_counts = Counter(t.get("post_type", "text") for t in tweets)

        # この人の話題
        topics = Counter()
        for t in tweets:
            for topic in detect_topics(t):
                topics[topic] += 1

        # テキストがある投稿のサンプル（内容把握用）
        text_samples = []
        for t in sorted(tweets, key=lambda x: x["metrics"]["likes"], reverse=True):
            text = t.get("text", "").strip()
            title = t.get("_title", "")
            if title:
                text_samples.append(f"「{title}」")
            elif not re.match(r'^https?://t\.co/\S+$', text) and len(text) > 20:
                text_samples.append(text[:80])
            if len(text_samples) >= 3:
                break

        profiles.append({
            "username": username,
            "followers": data["followers"],
            "count": len(tweets),
            "total_likes": total_likes,
            "total_bmarks": total_bmarks,
            "main_type": type_counts.most_common(1)[0][0],
            "topics": topics.most_common(3),
            "samples": text_samples,
            "account_url": data["account_url"],
        })

    return sorted(profiles, key=lambda x: x["total_likes"], reverse=True)


# ============================================================
# Markdown 生成
# ============================================================

def generate_md(name, all_tweets, per_label, labels, queries=None):
    lines = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    total = len(all_tweets)

    lines.append(f"# {name}")
    lines.append(f"")
    lines.append(f"> 生成日時: {now} | 合計: {total}件 | X直近7日間")
    lines.append(f"")

    if not all_tweets:
        lines.append("データなし。")
        return "\n".join(lines)

    # --- Pre-compute ---
    likes = [t["metrics"]["likes"] for t in all_tweets]
    bmarks = [t["metrics"].get("bookmarks", 0) for t in all_tweets]
    total_likes = sum(likes)
    avg_likes = total_likes / len(likes)
    max_t = max(all_tweets, key=lambda t: t["metrics"]["likes"])
    total_bmarks = sum(bmarks)
    save_rate = total_bmarks / total_likes if total_likes > 0 else 0
    type_counts = Counter(t.get("post_type", "text") for t in all_tweets)
    top10 = sorted(all_tweets, key=lambda t: t["metrics"]["likes"], reverse=True)[:10]
    eff_sorted = sorted(all_tweets, key=lambda t: t["metrics"]["likes"] / max(t.get("author_followers", 1), 1), reverse=True)
    save_sorted = sorted(
        [t for t in all_tweets if t["metrics"]["likes"] >= 50],
        key=lambda t: t["metrics"].get("bookmarks", 0) / max(t["metrics"]["likes"], 1),
        reverse=True
    )
    topic_map = analyze_topics(all_tweets)
    account_profiles = analyze_accounts(all_tweets)

    # X記事でテキストがURL-onlyかつタイトル未取得のもの
    untitled_articles = [
        t for t in all_tweets
        if t.get("post_type") == "x_article"
        and not t.get("_title")
        and re.match(r'^https?://t\.co/\S+$', t.get("text", "").strip())
    ]

    # === 何が語られているか ===
    lines.append(f"## 何が語られているか")
    lines.append(f"")
    if topic_map:
        used_example_ids = set()
        for topic, tweets in topic_map[:5]:
            topic_likes = sum(t["metrics"]["likes"] for t in tweets)
            # 既に例として使ったツイートを避けて選ぶ
            sorted_tweets = sorted(tweets, key=lambda t: t["metrics"]["likes"], reverse=True)
            top_tweet = next((t for t in sorted_tweets if t["id"] not in used_example_ids), None)
            if top_tweet is None:
                top_tweet = sorted_tweets[0]
                sample = get_display_text(top_tweet, max_len=80).replace("\n", " ") + "（再掲）"
            else:
                sample = get_display_text(top_tweet, max_len=80).replace("\n", " ")
            used_example_ids.add(top_tweet["id"])
            lines.append(f"- **{topic}**（{len(tweets)}件 / {compact(topic_likes)}いいね）— 例: {sample}")
        lines.append(f"")
    else:
        lines.append(f"テキストから話題を検出できませんでした。X記事が多い場合は `--titles` でタイトルを渡してください。")
        lines.append(f"")

    if untitled_articles:
        lines.append(f"> ⚠ X記事{len(untitled_articles)}件はAPIからタイトル取得不可。`--titles` でタイトルJSONを渡すと内容が反映されます。")
        lines.append(f"")

    # === キーパーソン ===
    lines.append(f"## キーパーソン")
    lines.append(f"")
    for p in account_profiles[:8]:
        if p["total_likes"] < 10: continue
        if not p["topics"]: continue
        pt_label = POST_TYPE_LABELS.get(p["main_type"], "?")
        topic_str = "、".join(t for t, _ in p["topics"]) if p["topics"] else "話題不明"
        lines.append(f"### @{p['username']}（{compact(p['followers'])}フォロワー / {p['count']}件 / 計{compact(p['total_likes'])}いいね）")
        lines.append(f"")
        lines.append(f"- **話題**: {topic_str} | **主な形式**: {pt_label}")
        if p["samples"]:
            for s in p["samples"][:3]:
                # 改行を除去して1行に
                s_clean = s.replace("\n", " ").strip()
                if len(s_clean) > 100:
                    s_clean = s_clean[:100] + "…"
                lines.append(f"- {s_clean}")
        lines.append(f"")

    # === アクションプラン ===
    lines.append(f"## 次にやるべきこと")
    lines.append(f"")

    # フォーマット戦略
    top10_types = Counter(t.get("post_type", "text") for t in top10)
    top10_best = top10_types.most_common(1)[0]
    top10_best_label = POST_TYPE_LABELS.get(top10_best[0], top10_best[0])
    lines.append(f"1. **フォーマット**: TOP10では「{top10_best_label}」が{top10_best[1]}/10件。")

    # 話題戦略
    if topic_map:
        best_topic = topic_map[0]
        best_topic_likes = sum(t["metrics"]["likes"] for t in best_topic[1])
        lines.append(f"2. **狙うべき話題**: 「{best_topic[0]}」が{compact(best_topic_likes)}いいねで最も反応が強い。")

    # ラベル比較
    if len(per_label) > 1:
        label_stats = {}
        for label, tweets in per_label.items():
            if not tweets: continue
            avg_l = sum(t["metrics"]["likes"] for t in tweets) / len(tweets)
            label_stats[label] = avg_l
        if label_stats:
            best = max(label_stats.items(), key=lambda x: x[1])
            lines.append(f"3. **切り口**: 「{best[0]}」が平均{compact(best[1])}いいねで最も強い。")

    # 保存率
    if save_rate >= 0.5:
        lines.append(f"4. **保存率{save_rate:.0%}**: 「後で見返したい」実用コンテンツの需要が高い。ハウツー系で出すのが効果的。")
    elif save_rate >= 0.3:
        lines.append(f"4. **保存率{save_rate:.0%}**: 実用的な情報への需要あり。")

    # 避けるべき
    bottom = sorted(all_tweets, key=lambda t: t["metrics"]["likes"])[:10]
    bottom_types = Counter(t.get("post_type", "text") for t in bottom)
    bottom_top = bottom_types.most_common(1)[0]
    bottom_label = POST_TYPE_LABELS.get(bottom_top[0], bottom_top[0])
    lines.append(f"5. **避けるべき**: いいね下位10件は「{bottom_label}」が{bottom_top[1]}/10件。")
    lines.append(f"")

    # === バズTOP10 ===
    lines.append(f"## バズTOP10")
    lines.append(f"")
    for i, t in enumerate(top10, 1):
        m = t["metrics"]
        tags = tag_buzz_reason(t)
        tag_str = " ".join(f"`{tag}`" for tag in tags)
        followers = t.get("author_followers", 1) or 1
        eff = m["likes"] / followers
        pt_label = POST_TYPE_LABELS.get(t.get("post_type", "text"), "?")
        display = get_display_text(t)
        # 改行を除去して読みやすく
        display_clean = display.replace("\n", " ").strip()
        if len(display_clean) > 200:
            display_clean = display_clean[:200] + "…"

        lines.append(f"**{i}. @{t['username']}** — {compact(m['likes'])}いいね / {compact(m.get('bookmarks',0))}ブクマ（{pt_label} / 効率{eff:.1f}x）")
        lines.append(f"")
        lines.append(f"> {display_clean}")
        lines.append(f"")
        lines.append(f"{tag_str} — [{t.get('tweet_url', '')}]({t.get('tweet_url', '')})")
        lines.append(f"")

    # === 数値サマリー ===
    lines.append(f"## 数値サマリー")
    lines.append(f"")

    # 検索クエリ
    lines.append(f"**検索クエリ:**")
    for i, (label, tweets) in enumerate(per_label.items()):
        q_str = ""
        if queries and i < len(queries):
            q_str = f" — `{queries[i]}`"
        lines.append(f"- {label}: {len(tweets)}件{q_str}")
    lines.append(f"")

    lines.append(f"| 指標 | 値 |")
    lines.append(f"|------|-----|")
    lines.append(f"| 投稿数 | {total}件 |")
    lines.append(f"| 合計いいね | {compact(total_likes)} |")
    lines.append(f"| 平均いいね | {compact(avg_likes)} |")
    lines.append(f"| 最大いいね | {compact(max_t['metrics']['likes'])} (@{max_t['username']}) |")
    lines.append(f"| 平均保存率 | {save_rate:.1%} |")
    lines.append(f"")

    if type_counts:
        type_str = " / ".join(f"{POST_TYPE_LABELS.get(pt, pt)}: {c}件" for pt, c in type_counts.most_common())
        lines.append(f"**投稿タイプ**: {type_str}")
        lines.append(f"")

    if len(per_label) > 1:
        lines.append(f"### ラベル別比較")
        lines.append(f"")
        lines.append(f"| ラベル | 件数 | 平均いいね | 最大 | 保存率 |")
        lines.append(f"|--------|------|-----------|------|--------|")
        for label, tweets in per_label.items():
            if not tweets: continue
            l = [t["metrics"]["likes"] for t in tweets]
            b = [t["metrics"].get("bookmarks", 0) for t in tweets]
            top = max(tweets, key=lambda t: t["metrics"]["likes"])
            sr = sum(b) / sum(l) if sum(l) > 0 else 0
            lines.append(f"| {label} | {len(tweets)} | {compact(sum(l)/len(l))} | {compact(max(l))} (@{top['username']}) | {sr:.1%} |")
        lines.append(f"")

    # === 保存されるコンテンツ ===
    if save_sorted:
        lines.append(f"## 保存されるコンテンツ（保存率TOP5）")
        lines.append(f"")
        for i, t in enumerate(save_sorted[:5], 1):
            sr = t["metrics"].get("bookmarks", 0) / t["metrics"]["likes"]
            pt_label = POST_TYPE_LABELS.get(t.get("post_type", "text"), "?")
            display = get_display_text(t, max_len=100)
            display_clean = display.replace("\n", " ")
            lines.append(f"{i}. **@{t['username']}** (保存率{sr:.0%} / {compact(t['metrics']['likes'])}L) — {display_clean}")
            lines.append(f"   [{t.get('tweet_url', '')}]({t.get('tweet_url', '')})")
            lines.append(f"")

    # === 外部リンク ===
    ext_urls = []
    for t in all_tweets:
        for um in t.get("url_meta", []):
            eu = um.get("expanded_url", "")
            title = um.get("title", "")
            if eu and "x.com" not in eu and "twitter.com" not in eu:
                ext_urls.append((eu, title, t["metrics"]["likes"], t["username"]))
    if ext_urls:
        ext_urls.sort(key=lambda x: x[2], reverse=True)
        seen_urls = set()
        lines.append(f"## 外部リンク")
        lines.append(f"")
        for url, title, lk, user in ext_urls[:10]:
            if url in seen_urls: continue
            seen_urls.add(url)
            label = title if title else url
            lines.append(f"- [{label}]({url}) — @{user}（{compact(lk)}いいね）")
        lines.append(f"")

    lines.append(f"---")
    lines.append(f"*Generated by x-research skill*")
    return "\n".join(lines)


# ============================================================
# xlsx 生成
# ============================================================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
NUM_FMT = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
PCT_FMT = '0.0%'

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)


def write_all_tweets_sheet(ws, all_tweets):
    ws.title = "全ツイート"
    headers = [
        "No", "ラベル", "ユーザー名", "フォロワー", "投稿タイプ", "話題",
        "テキスト", "いいね", "RT", "引用", "リプライ", "インプ", "ブクマ",
        "バズ効率", "保存率", "バズ要因タグ", "ポストURL", "アカウントURL",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    sorted_tweets = sorted(all_tweets, key=lambda t: t["metrics"]["likes"], reverse=True)
    for i, t in enumerate(sorted_tweets, 1):
        m = t["metrics"]
        followers = t.get("author_followers", 0) or 0
        eff = m["likes"] / max(followers, 1)
        sr = m.get("bookmarks", 0) / max(m["likes"], 1)
        pt = POST_TYPE_LABELS.get(t.get("post_type", "text"), "?")
        tags = ", ".join(tag_buzz_reason(t))
        topics = ", ".join(detect_topics(t)) or "—"
        display = get_display_text(t)

        ws.append([
            i, t.get("_label", ""), f"@{t.get('username', '?')}", followers, pt, topics,
            display, m["likes"], m.get("retweets", 0), m.get("quotes", 0),
            m.get("replies", 0), m.get("impressions", 0), m.get("bookmarks", 0),
            eff, sr, tags, t.get("tweet_url", ""), t.get("account_url", ""),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in [3, 7, 8, 9, 10, 11, 12]:
            row[col_idx].number_format = NUM_FMT
        row[13].number_format = '0.0x'
        row[14].number_format = PCT_FMT

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        try:
            if row[13].value and float(row[13].value) >= 1.0:
                for cell in row:
                    cell.fill = GREEN_FILL
        except (ValueError, TypeError):
            pass

    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.column_dimensions["G"].width = 60


def write_account_sheet(wb, all_tweets):
    ws = wb.create_sheet("アカウント別")
    headers = [
        "ユーザー名", "フォロワー", "投稿数", "合計いいね", "平均いいね",
        "合計ブクマ", "平均保存率", "主な投稿タイプ", "話題", "最大バズ", "アカウントURL",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    profiles = analyze_accounts(all_tweets)
    for p in profiles:
        pt_label = POST_TYPE_LABELS.get(p["main_type"], "?")
        topic_str = ", ".join(t for t, _ in p["topics"]) or "—"
        avg_sr = p["total_bmarks"] / max(p["total_likes"], 1)
        max_likes = max(t["metrics"]["likes"] for t in [tw for tw in all_tweets if tw.get("username") == p["username"]])
        ws.append([
            f"@{p['username']}", p["followers"], p["count"], p["total_likes"],
            p["total_likes"] / p["count"], p["total_bmarks"], avg_sr, pt_label,
            topic_str, max_likes, p["account_url"],
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[1].number_format = NUM_FMT
        row[3].number_format = NUM_FMT
        row[4].number_format = NUM_FMT
        row[5].number_format = NUM_FMT
        row[6].number_format = PCT_FMT
        row[9].number_format = NUM_FMT

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_label_sheet(wb, per_label):
    ws = wb.create_sheet("ラベル別")
    headers = [
        "ラベル", "件数", "合計いいね", "平均いいね", "最大いいね",
        "合計ブクマ", "保存率", "主な投稿タイプ", "トップユーザー",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for label, tweets in per_label.items():
        if not tweets: continue
        total_likes = sum(t["metrics"]["likes"] for t in tweets)
        total_bmarks = sum(t["metrics"].get("bookmarks", 0) for t in tweets)
        max_t = max(tweets, key=lambda t: t["metrics"]["likes"])
        type_counts = Counter(t.get("post_type", "text") for t in tweets)
        main_type = POST_TYPE_LABELS.get(type_counts.most_common(1)[0][0], "?")
        user_likes = Counter()
        for t in tweets:
            user_likes[t.get("username", "?")] += t["metrics"]["likes"]
        top_user = user_likes.most_common(1)[0][0] if user_likes else "?"
        ws.append([
            label, len(tweets), total_likes, total_likes / len(tweets),
            max_t["metrics"]["likes"], total_bmarks,
            total_bmarks / max(total_likes, 1), main_type, f"@{top_user}",
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[2].number_format = NUM_FMT
        row[3].number_format = NUM_FMT
        row[4].number_format = NUM_FMT
        row[5].number_format = NUM_FMT
        row[6].number_format = PCT_FMT

    auto_width(ws)


def write_type_sheet(wb, all_tweets):
    ws = wb.create_sheet("投稿タイプ別")
    headers = ["投稿タイプ", "件数", "合計いいね", "平均いいね", "合計ブクマ", "平均保存率", "平均バズ効率"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    by_type = defaultdict(list)
    for t in all_tweets:
        by_type[t.get("post_type", "text")].append(t)

    rows = []
    for pt, tweets in by_type.items():
        total_likes = sum(t["metrics"]["likes"] for t in tweets)
        total_bmarks = sum(t["metrics"].get("bookmarks", 0) for t in tweets)
        avg_eff = sum(t["metrics"]["likes"] / max(t.get("author_followers", 1), 1) for t in tweets) / len(tweets)
        rows.append([
            POST_TYPE_LABELS.get(pt, pt), len(tweets), total_likes,
            total_likes / len(tweets), total_bmarks,
            total_bmarks / max(total_likes, 1), avg_eff,
        ])

    rows.sort(key=lambda r: r[2], reverse=True)
    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[2].number_format = NUM_FMT
        row[3].number_format = NUM_FMT
        row[4].number_format = NUM_FMT
        row[5].number_format = PCT_FMT
        row[6].number_format = '0.00x'

    auto_width(ws)


def generate_xlsx(xlsx_path, all_tweets, per_label):
    wb = Workbook()
    write_all_tweets_sheet(wb.active, all_tweets)
    write_account_sheet(wb, all_tweets)
    if len(per_label) > 1:
        write_label_sheet(wb, per_label)
    write_type_sheet(wb, all_tweets)
    wb.save(str(xlsx_path))


# ============================================================
# main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="X Research → Markdown + xlsx バズ分析")
    parser.add_argument("--name", required=True, help="レポートのテーマ名")
    parser.add_argument("--files", nargs="+", required=True, help="JSONファイルのパス")
    parser.add_argument("--labels", nargs="+", help="各ファイルのラベル（省略時はファイル名）")
    parser.add_argument("--queries", nargs="+", help="各ファイルの検索クエリ文字列（省略可）")
    parser.add_argument("--titles", help="X記事タイトルのJSONマッピング（{tweet_id: title}）")
    parser.add_argument("--out-dir", default=str(Path.home() / ".claude/skills/x-research/reports"), help="出力ディレクトリ")
    parser.add_argument("--no-xlsx", action="store_true", help="xlsx出力をスキップ")
    parser.add_argument("--exclude", nargs="+", help="除外するツイートID")
    parser.add_argument("--topics", help="TOPIC_RULESのJSONファイル（省略時はデフォルトルール）")
    parser.add_argument("--no-noise-filter", action="store_true", help="自動ノイズ除去を無効化")
    args = parser.parse_args()

    labels = args.labels if args.labels and len(args.labels) == len(args.files) else [Path(f).stem for f in args.files]
    exclude_ids = set(args.exclude or [])

    # TOPIC_RULES差し替え
    global TOPIC_RULES
    if args.topics:
        custom = json.loads(Path(args.topics).read_text())
        TOPIC_RULES = [(r["name"], r["keywords"]) for r in custom]
        print(f"[カスタムTOPIC_RULES] {len(TOPIC_RULES)}カテゴリ読み込み", file=sys.stderr)

    # タイトルマッピング読み込み
    title_map = None
    if args.titles:
        title_map = json.loads(Path(args.titles).read_text())

    all_tweets, per_label = load_and_dedupe(
        args.files, labels, title_map, exclude_ids,
        auto_noise=not args.no_noise_filter,
    )
    md = generate_md(args.name, all_tweets, per_label, labels, queries=args.queries)

    slug = args.name.replace(" ", "-").replace("/", "-").lower()
    out_dir = Path(args.out_dir) / datetime.now().strftime("%Y-%m-%d") / slug
    out_dir.mkdir(parents=True, exist_ok=True)

    md_path = out_dir / f"{slug}.md"
    md_path.write_text(md, encoding="utf-8")
    print(f"Saved: {md_path}", file=sys.stderr)

    if not args.no_xlsx and all_tweets:
        xlsx_path = out_dir / f"{slug}.xlsx"
        generate_xlsx(xlsx_path, all_tweets, per_label)
        print(f"Saved: {xlsx_path}", file=sys.stderr)

    print(md)


if __name__ == "__main__":
    main()
