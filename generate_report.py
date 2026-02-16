#!/usr/bin/env python3
"""
Claude Code & OpenClaw X/Twitter Research Report Generator
- Reads cached search data
- Separates Japanese vs English tweets
- Generates strategic insights
- Exports comprehensive xlsx to x_skill directory
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...", file=sys.stderr)
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

# --- Language Detection ---
JAPANESE_PATTERN = re.compile(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF\u3400-\u4DBF]')

def is_japanese(text):
    """Detect if text contains significant Japanese characters."""
    jp_chars = len(JAPANESE_PATTERN.findall(text))
    total_chars = len(text.strip())
    if total_chars == 0:
        return False
    return jp_chars / total_chars > 0.1 or jp_chars >= 5

# --- Styling ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
INSIGHT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
BOLD = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_width(ws, min_w=8, max_w=60):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)

def compact(n):
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000: return f"{n/1_000:.1f}K"
    return str(n)

# --- Load Data ---
CACHE_DIR = Path.home() / "0_AI" / ".claude" / "skills" / "x-research" / "data" / "cache"
OUTPUT_DIR = Path.home() / "0_AI" / "x_skill"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

all_tweets = []
for f in CACHE_DIR.glob("*.json"):
    try:
        data = json.loads(f.read_text())
        tweets = data.get("tweets", [])
        all_tweets.extend(tweets)
    except Exception as e:
        print(f"Skipping {f.name}: {e}", file=sys.stderr)

# Deduplicate
seen = set()
unique_tweets = []
for t in all_tweets:
    if t["id"] not in seen:
        seen.add(t["id"])
        unique_tweets.append(t)
all_tweets = unique_tweets

print(f"Loaded {len(all_tweets)} unique tweets from cache", file=sys.stderr)

# Filter 100+ likes
viral_tweets = [t for t in all_tweets if t.get("metrics", {}).get("likes", 0) >= 100]
viral_tweets.sort(key=lambda t: t["metrics"]["likes"], reverse=True)
print(f"Viral tweets (100+ likes): {len(viral_tweets)}", file=sys.stderr)

# Separate by language
ja_tweets = [t for t in viral_tweets if is_japanese(t["text"])]
en_tweets = [t for t in viral_tweets if not is_japanese(t["text"])]

print(f"Japanese: {len(ja_tweets)}, English: {len(en_tweets)}", file=sys.stderr)

# --- Analysis Functions ---
def analyze_tweets(tweets):
    if not tweets:
        return {}
    likes = [t["metrics"]["likes"] for t in tweets]
    imps = [t["metrics"].get("impressions", 0) for t in tweets]
    rts = [t["metrics"]["retweets"] for t in tweets]

    total_eng = sum(likes) + sum(rts) + sum(t["metrics"]["replies"] for t in tweets)
    total_imp = sum(imps)

    # Content type
    text_only = [t for t in tweets if not t.get("urls") or len(t["urls"]) == 0]
    with_media = [t for t in tweets if any("photo" in u or "video" in u for u in t.get("urls", []))]
    with_links = [t for t in tweets if t.get("urls") and not any("photo" in u or "video" in u for u in t.get("urls", []))]

    # By hour
    by_hour = {}
    for t in tweets:
        try:
            h = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00")).hour
            if h not in by_hour:
                by_hour[h] = {"likes": [], "count": 0}
            by_hour[h]["likes"].append(t["metrics"]["likes"])
            by_hour[h]["count"] += 1
        except:
            pass

    # Top users
    user_stats = {}
    for t in tweets:
        u = t["username"]
        if u not in user_stats:
            user_stats[u] = {"name": t.get("name", "?"), "tweets": [], "total_likes": 0, "total_imps": 0}
        user_stats[u]["tweets"].append(t)
        user_stats[u]["total_likes"] += t["metrics"]["likes"]
        user_stats[u]["total_imps"] += t["metrics"].get("impressions", 0)

    return {
        "count": len(tweets),
        "total_likes": sum(likes),
        "avg_likes": sum(likes) / len(likes),
        "median_likes": sorted(likes)[len(likes)//2],
        "max_likes": max(likes),
        "total_imps": total_imp,
        "avg_imps": total_imp / len(imps) if imps else 0,
        "eng_rate": total_eng / total_imp if total_imp > 0 else 0,
        "text_only": len(text_only),
        "with_media": len(with_media),
        "with_links": len(with_links),
        "by_hour": by_hour,
        "user_stats": user_stats,
    }

# --- Strategic Insights Generation ---
def generate_insights(ja_stats, en_stats, ja_tweets, en_tweets):
    insights = []

    # 1. Content format analysis
    all_tw = ja_tweets + en_tweets
    media_tweets = [t for t in all_tw if any("photo" in u or "video" in u for u in t.get("urls", []))]
    text_tweets = [t for t in all_tw if not t.get("urls") or len(t["urls"]) == 0]

    avg_likes_media = sum(t["metrics"]["likes"] for t in media_tweets) / len(media_tweets) if media_tweets else 0
    avg_likes_text = sum(t["metrics"]["likes"] for t in text_tweets) / len(text_tweets) if text_tweets else 0

    if avg_likes_media > avg_likes_text:
        insights.append({
            "category": "コンテンツ形式",
            "insight": f"メディア付きポストが平均{compact(int(avg_likes_media))}いいねで、テキストのみ（{compact(int(avg_likes_text))}いいね）を上回る。画像・動画を必ず添付すべき。",
            "priority": "HIGH"
        })
    else:
        insights.append({
            "category": "コンテンツ形式",
            "insight": f"テキストのみのポストが平均{compact(int(avg_likes_text))}いいねで高パフォーマンス。核心を突く短文が刺さる。",
            "priority": "HIGH"
        })

    # 2. Theme pattern analysis
    themes = {
        "生産性革命": ["確定申告", "税理士", "全部終わった", "saved", "built", "shipped", "productivity"],
        "ユーモア・ネタ": ["飛びました", "bubble", "meme", "average", "sleeping", "but for"],
        "How-to/Tips": ["tips", "how to", "guide", "master", "best practice", "CLAUDE.md", "workflow"],
        "新機能・ニュース": ["announcing", "agent team", "opus 4.6", "hackathon", "update", "new"],
        "衝撃データ": ["4%", "commits", "10,000", "stars", "million", "never", "二度と"],
        "ツール・エコシステム": ["figma", "chrome", "extension", "built", "open source"],
    }

    theme_scores = {}
    for theme, keywords in themes.items():
        matching = []
        for t in all_tw:
            text_lower = t["text"].lower()
            if any(kw.lower() in text_lower for kw in keywords):
                matching.append(t)
        if matching:
            avg_l = sum(t["metrics"]["likes"] for t in matching) / len(matching)
            theme_scores[theme] = {"count": len(matching), "avg_likes": avg_l, "top": max(matching, key=lambda x: x["metrics"]["likes"])}

    sorted_themes = sorted(theme_scores.items(), key=lambda x: x[1]["avg_likes"], reverse=True)
    if sorted_themes:
        top_theme = sorted_themes[0]
        insights.append({
            "category": "バズるテーマ",
            "insight": f"最もエンゲージメントが高いテーマは「{top_theme[0]}」（平均{compact(int(top_theme[1]['avg_likes']))}いいね）。このテーマで投稿すると反応が取れやすい。",
            "priority": "HIGH"
        })

        for theme, data in sorted_themes[1:4]:
            insights.append({
                "category": "バズるテーマ",
                "insight": f"「{theme}」テーマ：{data['count']}件、平均{compact(int(data['avg_likes']))}いいね。代表例: @{data['top']['username']}の投稿",
                "priority": "MEDIUM"
            })

    # 3. Account patterns to emulate
    if ja_stats.get("user_stats"):
        ja_top_users = sorted(ja_stats["user_stats"].items(), key=lambda x: x[1]["total_likes"], reverse=True)[:5]
        for username, data in ja_top_users:
            avg = data["total_likes"] / len(data["tweets"])
            insights.append({
                "category": "真似すべきアカウント（日本語圏）",
                "insight": f"@{username}（{data['name']}）: {len(data['tweets'])}件で合計{compact(data['total_likes'])}いいね（平均{compact(int(avg))}）。この人の投稿スタイルを参考にする。",
                "priority": "HIGH"
            })

    if en_stats.get("user_stats"):
        en_top_users = sorted(en_stats["user_stats"].items(), key=lambda x: x[1]["total_likes"], reverse=True)[:5]
        for username, data in en_top_users:
            avg = data["total_likes"] / len(data["tweets"])
            insights.append({
                "category": "真似すべきアカウント（英語圏）",
                "insight": f"@{username}（{data['name']}）: {len(data['tweets'])}件で合計{compact(data['total_likes'])}いいね（平均{compact(int(avg))}）。この人の投稿スタイルを参考にする。",
                "priority": "HIGH"
            })

    # 4. Posting time insights
    all_hours = {}
    for t in all_tw:
        try:
            h = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00")).hour
            if h not in all_hours:
                all_hours[h] = []
            all_hours[h].append(t["metrics"]["likes"])
        except:
            pass

    if all_hours:
        best_hours = sorted(all_hours.items(), key=lambda x: sum(x[1])/len(x[1]), reverse=True)[:3]
        hours_str = ", ".join(f"{h}:00 UTC（JST {(h+9)%24}:00）" for h, _ in best_hours)
        insights.append({
            "category": "投稿時間帯",
            "insight": f"最もエンゲージメントが高い投稿時間帯: {hours_str}。この時間帯に投稿すべき。",
            "priority": "MEDIUM"
        })

    # 5. Post length analysis
    short_posts = [t for t in all_tw if len(t["text"]) < 100]
    medium_posts = [t for t in all_tw if 100 <= len(t["text"]) < 280]
    long_posts = [t for t in all_tw if len(t["text"]) >= 280]

    length_data = []
    if short_posts:
        length_data.append(("短文(<100字)", sum(t["metrics"]["likes"] for t in short_posts)/len(short_posts), len(short_posts)))
    if medium_posts:
        length_data.append(("中文(100-280字)", sum(t["metrics"]["likes"] for t in medium_posts)/len(medium_posts), len(medium_posts)))
    if long_posts:
        length_data.append(("長文(280字+)", sum(t["metrics"]["likes"] for t in long_posts)/len(long_posts), len(long_posts)))

    if length_data:
        best = max(length_data, key=lambda x: x[1])
        insights.append({
            "category": "文字数",
            "insight": f"最もバズるのは{best[0]}（平均{compact(int(best[1]))}いいね、{best[2]}件）。この長さを意識して書くべき。",
            "priority": "MEDIUM"
        })

    # 6. Japanese vs English engagement comparison
    if ja_stats.get("count") and en_stats.get("count"):
        ja_avg = ja_stats["avg_likes"]
        en_avg = en_stats["avg_likes"]
        insights.append({
            "category": "言語別エンゲージメント",
            "insight": f"日本語: 平均{compact(int(ja_avg))}いいね（{ja_stats['count']}件）、英語: 平均{compact(int(en_avg))}いいね（{en_stats['count']}件）。{'日本語圏の方がエンゲージメントが高い' if ja_avg > en_avg else '英語圏の方がリーチが広い'}。",
            "priority": "HIGH"
        })

    # 7. Viral post pattern
    top_10 = sorted(all_tw, key=lambda t: t["metrics"]["likes"], reverse=True)[:10]
    patterns = []
    for t in top_10:
        text = t["text"]
        if len(text) < 80:
            patterns.append("短文インパクト")
        if "?" in text or "？" in text:
            patterns.append("疑問形")
        if any(w in text.lower() for w in ["never", "二度と", "もう", "all", "every", "全部"]):
            patterns.append("断言型")
        if any(w in text.lower() for w in ["just", "今", "ちょうど", "breaking"]):
            patterns.append("速報型")
        if any(w in text.lower() for w in ["how to", "tips", "方法", "guide", "master"]):
            patterns.append("ハウツー型")

    from collections import Counter
    pattern_counts = Counter(patterns)
    if pattern_counts:
        top_patterns = pattern_counts.most_common(3)
        pat_str = "、".join(f"{p}({c}件)" for p, c in top_patterns)
        insights.append({
            "category": "トップ10のバズパターン",
            "insight": f"上位10投稿に共通するパターン: {pat_str}。この型を意識して投稿すべき。",
            "priority": "HIGH"
        })

    return insights

# --- Build Report ---
ja_stats = analyze_tweets(ja_tweets)
en_stats = analyze_tweets(en_tweets)
insights = generate_insights(ja_stats, en_stats, ja_tweets, en_tweets)

# --- Create Workbook ---
wb = Workbook()

# Sheet 1: Strategic Insights
ws_insights = wb.active
ws_insights.title = "戦略的インサイト"

ws_insights.cell(row=1, column=1, value="Claude Code X/Twitter バズ分析 - 戦略的インサイト").font = TITLE_FONT
ws_insights.cell(row=2, column=1, value=f"分析日: {datetime.now().strftime('%Y-%m-%d')} | 対象: {len(viral_tweets)}件（100+いいね）| 日本語: {len(ja_tweets)}件 | 英語: {len(en_tweets)}件")
ws_insights.merge_cells("A1:E1")
ws_insights.merge_cells("A2:E2")

row = 4
headers = ["優先度", "カテゴリ", "インサイト"]
for col, h in enumerate(headers, 1):
    ws_insights.cell(row=row, column=col, value=h)
style_header(ws_insights, row, len(headers))

for insight in insights:
    row += 1
    ws_insights.cell(row=row, column=1, value=insight["priority"])
    ws_insights.cell(row=row, column=2, value=insight["category"])
    ws_insights.cell(row=row, column=3, value=insight["insight"])

    if insight["priority"] == "HIGH":
        for c in range(1, 4):
            ws_insights.cell(row=row, column=c).fill = INSIGHT_FILL
            ws_insights.cell(row=row, column=c).font = BOLD

    ws_insights.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

ws_insights.column_dimensions["A"].width = 12
ws_insights.column_dimensions["B"].width = 30
ws_insights.column_dimensions["C"].width = 80
ws_insights.freeze_panes = "A5"

# Sheet 2: All Viral Tweets (sorted by likes)
ws_all = wb.create_sheet("全バズ投稿")
headers_all = ["No", "言語", "Username", "Name", "テキスト", "いいね", "RT", "リプ", "インプレッション", "ブックマーク", "投稿日時", "URL"]
ws_all.append(headers_all)
style_header(ws_all, 1, len(headers_all))

num_fmt = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
for i, t in enumerate(viral_tweets, 1):
    m = t["metrics"]
    lang = "JA" if is_japanese(t["text"]) else "EN"
    ws_all.append([
        i, lang,
        f"@{t['username']}", t.get("name", "?"),
        t["text"][:300],
        m["likes"], m["retweets"], m["replies"],
        m.get("impressions", 0), m.get("bookmarks", 0),
        t.get("created_at", ""),
        t.get("tweet_url", "")
    ])

    # Color code by language
    row_num = ws_all.max_row
    if lang == "JA":
        ws_all.cell(row=row_num, column=2).fill = GREEN_FILL
    else:
        ws_all.cell(row=row_num, column=2).fill = ACCENT_FILL

    for col_idx in [6, 7, 8, 9, 10]:
        ws_all.cell(row=row_num, column=col_idx).number_format = num_fmt

auto_width(ws_all, max_w=50)
ws_all.column_dimensions["E"].width = 60
ws_all.freeze_panes = "A2"

# Sheet 3: Japanese tweets
ws_ja = wb.create_sheet("日本語バズ投稿")
headers_ja = ["No", "Username", "Name", "テキスト", "いいね", "RT", "リプ", "インプレッション", "投稿日時", "URL"]
ws_ja.append(headers_ja)
style_header(ws_ja, 1, len(headers_ja))

for i, t in enumerate(ja_tweets, 1):
    m = t["metrics"]
    ws_ja.append([
        i, f"@{t['username']}", t.get("name", "?"),
        t["text"][:300],
        m["likes"], m["retweets"], m["replies"],
        m.get("impressions", 0),
        t.get("created_at", ""),
        t.get("tweet_url", "")
    ])
    for col_idx in [5, 6, 7, 8]:
        ws_ja.cell(row=ws_ja.max_row, column=col_idx).number_format = num_fmt

auto_width(ws_ja, max_w=50)
ws_ja.column_dimensions["D"].width = 60
ws_ja.freeze_panes = "A2"

# Sheet 4: English tweets
ws_en = wb.create_sheet("英語バズ投稿")
ws_en.append(headers_ja)  # Same headers
style_header(ws_en, 1, len(headers_ja))

for i, t in enumerate(en_tweets, 1):
    m = t["metrics"]
    ws_en.append([
        i, f"@{t['username']}", t.get("name", "?"),
        t["text"][:300],
        m["likes"], m["retweets"], m["replies"],
        m.get("impressions", 0),
        t.get("created_at", ""),
        t.get("tweet_url", "")
    ])
    for col_idx in [5, 6, 7, 8]:
        ws_en.cell(row=ws_en.max_row, column=col_idx).number_format = num_fmt

auto_width(ws_en, max_w=50)
ws_en.column_dimensions["D"].width = 60
ws_en.freeze_panes = "A2"

# Sheet 5: Theme Analysis
ws_themes = wb.create_sheet("テーマ分析")
ws_themes.cell(row=1, column=1, value="テーマ別バズ分析").font = TITLE_FONT
ws_themes.merge_cells("A1:D1")

themes = {
    "生産性革命": ["確定申告", "税理士", "全部終わった", "saved", "built", "shipped", "productivity", "浮いた"],
    "ユーモア・ミーム": ["飛びました", "bubble", "meme", "average", "sleeping", "but for", "放置"],
    "How-to/Tips集": ["tips", "how to", "guide", "master", "best practice", "CLAUDE.md", "workflow", "方法"],
    "新機能・公式発表": ["announcing", "agent team", "opus 4.6", "hackathon", "update", "new feature", "エージェントチーム"],
    "衝撃データ・統計": ["4%", "commits", "10,000", "stars", "million", "never", "二度と", "90%"],
    "ツール連携": ["figma", "chrome", "extension", "built", "open source", "dexter", "openclaw"],
    "感情・意見": ["weird time", "sadness", "wonder", "crazy", "amazing", "すごい", "やばい", "神"],
}

row = 3
ws_themes.append(["テーマ", "ヒット数", "平均いいね", "代表ポスト"])
style_header(ws_themes, row, 4)

all_tw = viral_tweets
for theme, keywords in themes.items():
    matching = [t for t in all_tw if any(kw.lower() in t["text"].lower() for kw in keywords)]
    if matching:
        avg_l = sum(t["metrics"]["likes"] for t in matching) / len(matching)
        top = max(matching, key=lambda x: x["metrics"]["likes"])
        row += 1
        ws_themes.cell(row=row, column=1, value=theme)
        ws_themes.cell(row=row, column=2, value=len(matching))
        ws_themes.cell(row=row, column=3, value=int(avg_l))
        ws_themes.cell(row=row, column=4, value=f"@{top['username']}: {top['text'][:100]}...")
        ws_themes.cell(row=row, column=3).number_format = num_fmt
        ws_themes.cell(row=row, column=4).alignment = Alignment(wrap_text=True)

auto_width(ws_themes, max_w=50)
ws_themes.column_dimensions["D"].width = 80

# Sheet 6: User Rankings
ws_users = wb.create_sheet("アカウントランキング")
ws_users.cell(row=1, column=1, value="バズアカウントランキング（真似すべきアカウント）").font = TITLE_FONT
ws_users.merge_cells("A1:F1")

row = 3
ws_users.append(["Username", "Name", "投稿数", "合計いいね", "平均いいね", "トップ投稿"])
style_header(ws_users, row, 6)

user_map = {}
for t in viral_tweets:
    u = t["username"]
    if u not in user_map:
        user_map[u] = {"name": t.get("name", "?"), "tweets": [], "total_likes": 0}
    user_map[u]["tweets"].append(t)
    user_map[u]["total_likes"] += t["metrics"]["likes"]

sorted_users = sorted(user_map.items(), key=lambda x: x[1]["total_likes"], reverse=True)
for username, data in sorted_users[:30]:
    row += 1
    avg_l = data["total_likes"] / len(data["tweets"])
    top = max(data["tweets"], key=lambda x: x["metrics"]["likes"])
    ws_users.cell(row=row, column=1, value=f"@{username}")
    ws_users.cell(row=row, column=2, value=data["name"])
    ws_users.cell(row=row, column=3, value=len(data["tweets"]))
    ws_users.cell(row=row, column=4, value=data["total_likes"])
    ws_users.cell(row=row, column=5, value=int(avg_l))
    ws_users.cell(row=row, column=6, value=f"{top['text'][:120]}...")

    ws_users.cell(row=row, column=4).number_format = num_fmt
    ws_users.cell(row=row, column=5).number_format = num_fmt
    ws_users.cell(row=row, column=6).alignment = Alignment(wrap_text=True)

    # Highlight top 5
    if row <= 8:
        for c in range(1, 7):
            ws_users.cell(row=row, column=c).fill = INSIGHT_FILL

auto_width(ws_users, max_w=50)
ws_users.column_dimensions["F"].width = 80
ws_users.freeze_panes = "A4"

# --- Save ---
output_path = OUTPUT_DIR / f"claude-code-buzz-analysis-{datetime.now().strftime('%Y%m%d')}.xlsx"
wb.save(str(output_path))
print(f"\nSaved: {output_path}")
print(f"\nSheets: 戦略的インサイト | 全バズ投稿 | 日本語バズ投稿 | 英語バズ投稿 | テーマ分析 | アカウントランキング")
